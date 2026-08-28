"""Helpers partagés pour générer les exports Excel des rapports (incendie et
extincteurs) — mise en page minimale mais cohérente entre les deux."""
import io
from urllib.parse import quote

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

NAVY = "0A0B0D"
RED = "E11324"


def excel_workbook() -> Workbook:
    return Workbook()


def excel_entete_rapport(ws, *, organisation_nom: str, adresse: str, date_insp: str, statut: str, techniciens: str) -> None:
    """Bloc d'en-tête commun (adresse, date, statut, techniciens) en haut de la feuille."""
    ws["A1"] = f"{organisation_nom} — Export Excel"
    ws["A1"].font = Font(bold=True, size=14, color=RED)
    lignes = [
        ("Adresse", adresse),
        ("Date d'inspection", date_insp),
        ("Statut", statut),
        ("Technicien(s)", techniciens),
    ]
    for i, (label, valeur) in enumerate(lignes, start=2):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True, color=NAVY)
        ws.cell(row=i, column=2, value=valeur)


def excel_ligne_entetes(ws, colonnes: list[str], ligne: int) -> int:
    """Écrit la ligne d'en-têtes de tableau à la rangée donnée, stylée. Retourne cette rangée.

    Ne fige pas les volets : cette fonction est appelée plusieurs fois pour un
    même rapport (une section par tableau) — figer à chaque appel finit par
    figer presque toute la feuille sur les exports à sections multiples."""
    for col, titre in enumerate(colonnes, start=1):
        cell = ws.cell(row=ligne, column=col, value=titre)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return ligne


def excel_ajuster_largeurs(ws, colonnes: list[str], largeur_min: int = 12) -> None:
    for i, titre in enumerate(colonnes, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(largeur_min, len(titre) + 2)


def excel_nom_fichier(prefixe: str, adresse: str) -> str:
    """Nom de fichier lisible : « Incendie - 1450 Rue St-Laurent, Montréal.xlsx »
    — l'adresse inspectée identifie mieux le fichier que l'ID interne du rapport."""
    propre = "".join(c for c in adresse if c not in '\\/:*?"<>|').strip()
    return f"{prefixe} - {propre}.xlsx"


def excel_reponse(wb: Workbook, nom_fichier: str) -> HttpResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    # Repli ASCII (accents translittérés) + filename* encodé RFC 5987, pour que
    # les caractères accentués (é, à…) passent correctement dans l'en-tête HTTP.
    repli_ascii = nom_fichier.encode("ascii", "replace").decode("ascii")
    response["Content-Disposition"] = (
        f'attachment; filename="{repli_ascii}"; filename*=UTF-8\'\'{quote(nom_fichier)}'
    )
    return response
