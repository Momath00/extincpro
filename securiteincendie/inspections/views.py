from collections import Counter
from datetime import date

from django.http import HttpResponse
from rest_framework import permissions, status, viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView

from securiteincendie.emailing import logo_data_uri, organisation_logo_content

from accounts.models import Utilisateur
from openpyxl.styles import Font

from .excel_utils import (
    NAVY,
    RED,
    excel_ajuster_largeurs,
    excel_entete_rapport,
    excel_ligne_entetes,
    excel_nom_fichier,
    excel_reponse,
    excel_workbook,
)
from .models import (
    AppelService,
    Batiment,
    BoyauItem,
    Certificat,
    CertificatExtincteur,
    Client,
    Dispositif,
    EclairageUrgenceItem,
    ExtincteurItem,
    FicheE1,
    FicheE2,
    FicheLegende,
    Rapport,
    RapportEclairageUrgence,
    RapportExtincteur,
    SectionDispositif,
)
from .serializers import (
    AppelServiceCreateSerializer,
    AppelServiceDetailSerializer,
    AppelServiceListSerializer,
    BatimentSerializer,
    BoyauItemSerializer,
    ClientSerializer,
    DispositifSerializer,
    EclairageUrgenceItemSerializer,
    ExtincteurItemSerializer,
    FicheE1Serializer,
    FicheE2Serializer,
    FicheLegendeSerializer,
    HistoriqueRapportEclairageUrgenceSerializer,
    HistoriqueRapportExtincteurSerializer,
    HistoriqueRapportSerializer,
    RapportCreateSerializer,
    RapportDetailSerializer,
    RapportEclairageUrgenceCreateSerializer,
    RapportEclairageUrgenceDetailSerializer,
    RapportEclairageUrgenceListSerializer,
    RapportExtincteurCreateSerializer,
    RapportExtincteurDetailSerializer,
    RapportExtincteurListSerializer,
    RapportListSerializer,
    SectionDispositifSerializer,
)

# ── Titres des sections E2 ───────────────────────────────────────────────
E2_TITRES = {
    "e2_1":  "E2.1 — Essai du poste de contrôle principal",
    "e2_2":  "E2.2 — Système de recherche de personnes / téléphones d'urgence",
    "e2_3":  "E2.3 — Vérification du poste de contrôle",
    "e2_4":  "E2.4 — Alimentation principale (C.A.)",
    "e2_5":  "E2.5 — Alimentation de secours (batterie)",
    "e2_6":  "E2.6 — Répondeur / panneau d'annonce — type I",
    "e2_7":  "E2.7 — Répondeur / panneau d'annonce — type II",
    "e2_8":  "E2.8 — Indicateur d'alarme à distance",
    "e2_9":  "E2.9 — Imprimante",
    "e2_10": "E2.10 — Liaisons de données",
    "e2_11": "E2.11 — Dispositifs auxiliaires",
    "e2_12": "E2.12 — Remarques générales",
}

# ── Légende des dispositifs (référence des abréviations E3) ─────────────
LEGENDE_DISPOSITIFS = [
    ("PAI", "Panneau annonciateur d'alarme"),
    ("M", "Station manuelle"),
    ("S", "Détecteur de fumée"),
    ("C", "Cloche"),
    ("K", "Klaxon"),
    ("RHT", "Détecteur de chaleur"),
    ("FDL", "Résistance de fin de ligne"),
    ("PZ", "Piézo"),
    ("ISO", "Module isolateur"),
    ("ANN", "Panneau annonciateur d'alarme"),
    ("DFG", "Détecteur de fumée gaine ventilation"),
    ("TEL", "Téléphone d'urgence (pompier)"),
    ("IDG", "Gicleur débit"),
    ("IVG", "Interrupteur vanne gicleur"),
    ("IHP", "Interrupteur haute pression"),
    ("IBH", "Interrupteur de basse pression"),
    ("K/S", "Klaxon strobe"),
    ("MA", "Module adressable"),
]


# ── Légende du rapport extincteurs portatifs ─────────────────────────────
LEGENDE_EXTINCTEURS = [
    ("HT", "Test hydro, pour boyaux et/ou extincteurs, voir (Notes)"),
    ("T/O", "Les extincteurs ou les boyaux ont dépassé le temps recommandé, voir (Notes)"),
    ("MQ", "Extincteur ou boyaux manquant, doit être ajouté, voir (Notes)"),
    ("RM", "Recommandation, voir (Notes)"),
    ("D", "Déficience, voir (Notes)"),
    ("MT", "Maintenance requise, voir (Notes)"),
]


def _val_oui_non(v):
    if v is True:  return '<span style="color:#0d6b4f;font-weight:700;">Oui</span>'
    if v is False: return '<span style="color:#e11324;font-weight:700;">Non</span>'
    return '<span style="color:#9ca3af;">—</span>'


def _val_so(v):
    if v == "oui":        return '<span style="color:#0d6b4f;font-weight:700;">Oui</span>'
    if v == "non":        return '<span style="color:#e11324;font-weight:700;">Non</span>'
    if v == "sans_objet": return '<span style="color:#6b7280;">S.O.</span>'
    if v:                 return f'<span style="color:#0a0b0d;">{v}</span>'
    return '<span style="color:#9ca3af;">—</span>'


_MOIS_FR = ['janvier','février','mars','avril','mai','juin','juillet','août',
            'septembre','octobre','novembre','décembre']

def _date_fr(d):
    if d is None: return "—"
    return f"{d.day} {_MOIS_FR[d.month - 1]} {d.year}"


def _creer_rapport_eclairage_lie(rapport_extincteur, utilisateur):
    """Crée le rapport de vérification de l'éclairage d'urgence lié à ce
    rapport extincteur — même visite, un seul certificat unifié à la
    fermeture. N'y touche pas si l'organisation n'a pas activé ce module.
    Appelé après la création d'un RapportExtincteur, qu'il soit créé
    directement via l'API ou automatiquement avec le rapport principal."""
    organisation = getattr(utilisateur, "organisation", None)
    if not (organisation and organisation.a_le_module("rapport_eclairage_urgence")):
        return

    rapport_eclairage = RapportEclairageUrgence.objects.create(
        batiment=rapport_extincteur.batiment,
        cree_par=utilisateur,
        numero_job=rapport_extincteur.numero_job,
        date_inspection=rapport_extincteur.date_inspection,
        rapport_extincteur=rapport_extincteur,
    )
    rapport_eclairage.techniciens.set(rapport_extincteur.techniciens.all())
    rapport_eclairage.historiser(
        utilisateur, "Rapport créé automatiquement avec le rapport extincteur"
    )


def _est_conforme_extincteur(rapport_extincteur):
    """Non conforme dès qu'un extincteur OU une unité d'éclairage d'urgence
    liée est défectueux — même logique que le certificat unifié."""
    rapport_eclairage = getattr(rapport_extincteur, "rapport_eclairage_lie", None)
    eclairages = list(rapport_eclairage.eclairages_urgence.all()) if rapport_eclairage else []
    items = list(rapport_extincteur.extincteurs.all())
    return not any(it.etat == "D" for it in items) and not any(it.etat == "D" for it in eclairages)


# ── Traduction des rapports/certificats (langue de l'organisation) ─────────
# Chaque organisation choisit sa langue (super-admin) — les documents générés
# (rapports imprimables, certificats) suivent cette langue, indépendamment de
# la langue de l'utilisateur qui les télécharge.
I18N = {
    "imprimer_pdf": {"fr": "Imprimer / Enregistrer PDF", "en": "Print / Save as PDF"},
    "certificat_verification": {"fr": "Certificat de vérification", "en": "Verification Certificate"},
    "extincteurs_portatifs": {"fr": "Extincteurs portatifs", "en": "Portable fire extinguishers"},
    "inspection_certification": {"fr": "Inspection &amp; Certification", "en": "Inspection &amp; Certification"},
    "date_inspection": {"fr": "Date d'inspection", "en": "Inspection date"},
    "technicien_s": {"fr": "Technicien(s)", "en": "Technician(s)"},
    "client": {"fr": "Client", "en": "Client"},
    "adresse_inspectee": {"fr": "Adresse inspectée", "en": "Inspected address"},
    "conformite_bandeau": {
        "fr": "La vérification de l'équipement sous mentionné est conforme aux normes en vigueur",
        "en": "The verification of the equipment listed below complies with applicable standards",
    },
    "equipement": {"fr": "Équipement", "en": "Equipment"},
    "conforme_col": {"fr": "Conforme", "en": "Compliant"},
    "non_conforme_col": {"fr": "Non conforme", "en": "Non-compliant"},
    "so": {"fr": "S.O.", "en": "N/A"},
    "statut_col": {"fr": "Statut", "en": "Status"},
    "systeme_cuisine": {"fr": "Système automatique de cuisine", "en": "Automatic kitchen system"},
    "extincteur_label": {"fr": "Extincteur", "en": "Fire extinguisher"},
    "eclairage_urgence_label": {"fr": "Éclairage d'urgence", "en": "Emergency lighting"},
    "conforme_badge": {"fr": "CONFORME", "en": "COMPLIANT"},
    "non_conforme_badge": {"fr": "NON CONFORME", "en": "NON-COMPLIANT"},
    "inspection_entretien": {
        "fr": "L'inspection régulière et l'entretien de l'équipement tels que recommandés<br>par le manufacturier ont été effectués.",
        "en": "Regular inspection and maintenance of the equipment as recommended<br>by the manufacturer have been carried out.",
    },
    "superviseur_responsable": {"fr": "Superviseur / Responsable", "en": "Supervisor / Responsible"},
    "date_emission": {"fr": "Date d'émission", "en": "Issue date"},
    "certificat_no": {"fr": "Certificat N°", "en": "Certificate No."},
    "footer_certificat_extincteur": {
        "fr": "Ce certificat atteste la vérification des extincteurs portatifs et de l'éclairage d'urgence à la date d'inspection indiquée.",
        "en": "This certificate attests to the verification of portable fire extinguishers and emergency lighting on the inspection date indicated.",
    },
    "footer_certificat_incendie": {
        "fr": "Ce certificat atteste la conformité du réseau d'alarme incendie à la date d'inspection indiquée.",
        "en": "This certificate attests to the compliance of the fire alarm system on the inspection date indicated.",
    },
    "footer_rapport_incendie": {
        "fr": "Ce rapport présente le détail de l'inspection annuelle du réseau d'alarme incendie à la date indiquée.",
        "en": "This report presents the details of the fire alarm system's annual inspection on the date indicated.",
    },
    "footer_rapport_extincteur": {
        "fr": "Ce rapport présente le détail de la vérification des extincteurs portatifs et de l'éclairage d'urgence à la date indiquée.",
        "en": "This report presents the details of the portable fire extinguisher and emergency lighting verification on the date indicated.",
    },
    "rapport_verification": {"fr": "Rapport de vérification", "en": "Verification Report"},
    "detail_extincteurs": {"fr": "Détail des extincteurs", "en": "Fire extinguisher details"},
    "detail_boyaux": {"fr": "Détail des boyaux d'incendie", "en": "Fire hose details"},
    "detail_unites_eclairage": {"fr": "Détail des unités d'éclairage d'urgence", "en": "Emergency lighting unit details"},
    "col_no": {"fr": "No", "en": "No."},
    "col_etage": {"fr": "Étage", "en": "Floor"},
    "col_emplacement": {"fr": "Emplacement", "en": "Location"},
    "col_type": {"fr": "Type", "en": "Type"},
    "col_format": {"fr": "Format", "en": "Size"},
    "col_marque": {"fr": "Marque", "en": "Brand"},
    "col_modele": {"fr": "Modèle", "en": "Model"},
    "col_voltage": {"fr": "Voltage", "en": "Voltage"},
    "col_numero_serie": {"fr": "N° série", "en": "Serial No."},
    "col_date_fabrication": {"fr": "Date fabrication", "en": "Manufacture date"},
    "col_prochaine_maintenance": {"fr": "Prochaine maintenance", "en": "Next maintenance"},
    "col_prochain_test_hydro": {"fr": "Prochain test hydro.", "en": "Next hydro test"},
    "col_longueur": {"fr": "Longueur", "en": "Length"},
    "col_annee_fabrication": {"fr": "Année fabrication", "en": "Manufacture year"},
    "col_etat": {"fr": "État", "en": "Status"},
    "col_remarque": {"fr": "Remarque", "en": "Remark"},
    "etat_titre_abbr": {
        "fr": "D=Défectueux, C=Conforme, NI=Non inspecté",
        "en": "D=Defective, C=Compliant, NI=Not inspected",
    },
    "adresse": {"fr": "Adresse", "en": "Address"},
    "statut_ferme": {"fr": "Fermé", "en": "Closed"},
    "statut_ouvert": {"fr": "Ouvert", "en": "Open"},
    "aucun_extincteur": {"fr": "Aucun extincteur enregistré", "en": "No fire extinguisher recorded"},
    "aucun_boyau": {"fr": "Aucun boyau enregistré", "en": "No fire hose recorded"},
    "aucune_unite": {"fr": "Aucune unité enregistrée", "en": "No unit recorded"},
    "aucun_appareil": {"fr": "Aucun appareil enregistré", "en": "No device recorded"},
    "footer_rapport_eclairage": {
        "fr": "Rapport de vérification — Éclairage d'urgence",
        "en": "Verification Report — Emergency Lighting",
    },
    "certificat_unifie_genere": {
        "fr": "généré depuis le rapport extincteur lié",
        "en": "generated from the linked fire extinguisher report",
    },
    "certificat_unifie_label": {"fr": "Certificat unifié", "en": "Unified certificate"},
    "total": {"fr": "Total", "en": "Total"},
    "defectueuse_s": {"fr": "défectueuse(s)", "en": "defective"},
    "boyaux_incendie_sheet": {"fr": "Boyaux", "en": "Hoses"},
}


def _t(langue, cle):
    return I18N[cle].get(langue, I18N[cle]["fr"])


# ── Traduction des valeurs de menus déroulants (format, type, marque,
# longueur) — distinctes des libellés d'interface : ce sont les valeurs de
# données choisies par le technicien, affichées dans les documents générés.
CHOIX_I18N = {
    "format": {
        "2.5lb": {"fr": "2.5 lb", "en": "2.5 lb"}, "5lb": {"fr": "5 lb", "en": "5 lb"},
        "10lb": {"fr": "10 lb", "en": "10 lb"}, "13.25lb": {"fr": "13.25 lb", "en": "13.25 lb"},
        "20lb": {"fr": "20 lb", "en": "20 lb"}, "2.5kg": {"fr": "2.5 kg", "en": "2.5 kg"},
        "5kg": {"fr": "5 kg", "en": "5 kg"}, "10kg": {"fr": "10 kg", "en": "10 kg"},
        "6L": {"fr": "6 L", "en": "6 L"}, "autre": {"fr": "Autre", "en": "Other"},
    },
    "type_extincteur": {
        "ABC": {"fr": "Poudre ABC", "en": "ABC Powder"}, "BC": {"fr": "Poudre BC", "en": "BC Powder"},
        "CO2": {"fr": "CO2", "en": "CO2"}, "EAU": {"fr": "Eau", "en": "Water"},
        "AFFF": {"fr": "Mousse (AFFF)", "en": "Foam (AFFF)"},
        "K": {"fr": "Produits chimiques humides (K)", "en": "Wet chemical (K)"},
        "halotron": {"fr": "Halotron", "en": "Halotron"}, "fe36": {"fr": "FE36", "en": "FE36"},
        "autre": {"fr": "Autre", "en": "Other"},
    },
    "marque": {
        "amerex": {"fr": "Amerex", "en": "Amerex"}, "kidde": {"fr": "Kidde", "en": "Kidde"},
        "buckeye": {"fr": "Buckeye", "en": "Buckeye"}, "ansul": {"fr": "Ansul", "en": "Ansul"},
        "general": {"fr": "General", "en": "General"}, "flag": {"fr": "Flag", "en": "Flag"},
        "strikefirst": {"fr": "Strike First", "en": "Strike First"}, "autre": {"fr": "Autre", "en": "Other"},
    },
    "longueur": {
        "50pi": {"fr": "50 pi", "en": "50 ft"}, "75pi": {"fr": "75 pi", "en": "75 ft"},
        "100pi": {"fr": "100 pi", "en": "100 ft"}, "autre": {"fr": "Autre", "en": "Other"},
    },
}


def _td(langue, categorie, code):
    """Traduit une valeur de choix de données (ex. marque, format) — retourne
    le code tel quel si non trouvé, pour ne jamais planter sur une valeur
    inattendue."""
    if not code:
        return code
    entree = CHOIX_I18N.get(categorie, {}).get(code)
    if not entree:
        return code
    return entree.get(langue, entree["fr"])


# ── Labels complets E2 ────────────────────────────────────────────────────
E2_ITEMS = {
    "e2_1": [
        ("A",  "Fonctionnement de l'indicateur visuel de mise sous tension"),
        ("B",  "Fonctionnement du signal de défectuosité visuel commun"),
        ("C",  "Fonctionnement du signal de défectuosité sonore commun"),
        ("D",  "Fonctionnement de l'interrupteur de signalisation sonore de défectuosité"),
        ("E",  "Fonctionnement du signal de défectuosité de l'alimentation principale"),
        ("F",  "Fuite à la terre sur signal de défectuosité positif et négatif"),
        ("G",  "Fonctionnement du signal d'alerte"),
        ("H",  "Fonctionnement du signal d'alarme"),
        ("I",  "Fonctionnement du passage automatique de signal d'alerte à signal d'alarme"),
        ("J",  "Fonctionnement du passage manuel de signal d'alerte à signal d'alarme"),
        ("K",  "Caractéristique d'annulation du passage automatique de signal d'alerte à signal d'alarme fonctionnant sur un réseau à deux étapes"),
        ("L",  "Fonctionnement de la désactivation de l'interruption du signal d'alarme sonore"),
        ("M",  "Fonctionnement de l'interruption manuelle du signal d'alarme sonore"),
        ("N",  "Fonctionnement de l'indicateur visuel d'interruption du signal d'alarme sonore"),
        ("O",  "Déclenchement automatique du signal d'alarme sonore, après interruption, en cas de réception d'alarme subséquente"),
        ("P",  "Temporisation automatique d'annulation du signal d'alarme sonore"),
        ("Q",  "Signaux d'alerte et d'alarme sonores et visuels programmés et fonctionnant conformément à la conception et aux spécifications"),
        ("R",  "Fonctionnement d'alarme et de surveillance du circuit d'entrée, y compris les indications sonores et visuelles"),
        ("S",  "La surveillance des défauts sur un circuit d'entrée entraîne une indication de défectuosité"),
        ("T",  "Fonctionnement des indicateurs d'alarme du circuit de sortie"),
        ("U",  "La surveillance des défauts sur un circuit de sortie entraîne une indication de défectuosité"),
        ("V",  "Essai d'indicateur visuel (essai de lampe)"),
        ("W",  "Séquences de signal codé fonctionnant au moins le nombre de fois nécessaire et suivies d'un déclenchement de signal d'alarme approprié"),
        ("X",  "Séquences de signal codé non interrompues par une alarme subséquente"),
        ("Y",  "Une dérivation du dispositif auxiliaire provoque un signal de défectuosité"),
        ("Z",  "Fonctionnement du circuit d'entrée vers le circuit de sortie, y compris les circuits des dispositifs auxiliaires, pour assurer le bon fonctionnement du programme"),
        ("AA", "Fonctionnement du réarmement du réseau avertisseur d'incendie"),
        ("BB", "Fonctionnement de la commutation de l'alimentation principale à l'alimentation de secours"),
        ("CC", "Vérification de la confirmation du changement d'état (détecteurs de fumée seulement ; se reporter au paragraphe 5.4.7.3)"),
        ("DD", "Réception de la transmission d'un signal d'alarme à la centrale de réception d'alarme incendie"),
        ("EE", "Réception de la transmission d'un signal de surveillance à la centrale de réception d'alarme incendie"),
        ("FF", "Réception de la transmission d'un signal de défectuosité à la centrale de réception d'alarme incendie"),
        ("GG", "Nom et numéro de téléphone de la centrale de réception d'alarme incendie"),
        ("HH", "Le déclenchement du sectionneur de la centrale de réception d'alarme incendie produit une indication de défectuosité précise au poste de contrôle et achemine un signal de défectuosité à la centrale"),
    ],
    "e2_2": [
        ("A",  "Fonctionnement de l'indicateur de mise sous tension"),
        ("B",  "Fonctionnement du signal de défectuosité visuel commun"),
        ("C",  "Fonctionnement du signal de défectuosité sonore commun"),
        ("D",  "Fonctionnement de l'interrupteur de signalisation sonore de défectuosité"),
        ("E",  "Fonctionnement de la recherche phonique générale de personnes, y compris l'indication visuelle"),
        ("F",  "Fonctionnement des circuits de sortie en cas de recherche phonique sélective de personnes, y compris l'indication visuelle"),
        ("G",  "Fonctionnement des circuits de sortie pour défectuosité de recherche phonique sélective de personnes, y compris l'indication visuelle"),
        ("H",  "Fonctionnement du microphone, y compris bouton de communication"),
        ("I",  "Fonctionnement de la recherche de personnes ne nuisant pas à la temporisation initiale de désactivation de la signalisation sonore d'alerte et d'alarme"),
        ("J",  "Fonctionnement de la recherche générale de personnes"),
        ("K",  "Passage automatique à un amplificateur de relève en cas de panne d'un amplificateur normal"),
        ("L",  "Circuits de réception d'appel d'un téléphone d'urgence, y compris les indications sonores et visuelles"),
        ("M",  "Fonctionnement des circuits des téléphones d'urgence, y compris les communications phoniques bidirectionnelles"),
        ("N",  "Fonctionnement des circuits de signalisation de défectuosité des téléphones d'urgence, y compris l'indication visuelle"),
        ("O",  "Fonctionnement des communications verbales par téléphone d'urgence"),
        ("P",  "Fonctionnement de la tonalité d'utilisation ou de disponibilité des téléphones d'urgence, au combiné"),
    ],
    "e2_3": [
        ("A",  "Désignations du circuit d'entrée correctement indiquées et correspondant aux dispositifs raccordés"),
        ("B",  "Désignations du circuit de sortie correctement repérées et correspondant à celles des dispositifs raccordés"),
        ("C",  "Désignations des fonctions de contrôle communes et des indicateurs communs correctes"),
        ("D",  "Composants enfichables et modules solidement en place"),
        ("E",  "Câbles enfichables solidement en place"),
        ("F",  "Date, version et révision des microprogrammes et des programmes logiciels consignés"),
        ("G",  "Propre et exempt de poussière et de saleté"),
        ("H",  "Fusibles conformes aux spécifications des fabricants"),
        ("I",  "Verrouillage du poste de contrôle ou du répondeur"),
        ("J",  "Solidité des connexions du câblage aux dispositifs"),
    ],
    "e2_4": [
        ("A",  "Protection fusible correspondant aux caractéristiques nominales affichées par le fabricant"),
        ("B",  "Alimentation suffisante pour les besoins du réseau"),
    ],
    "e2_5": [
        ("A",  "Type de batterie recommandée par le fabricant"),
        ("B",  "Caractéristiques nominales suffisantes après des calculs fondés sur la pleine charge du réseau"),
        ("C",  "Tension de batterie lorsque la source d'alimentation principale est sous tension"),
        ("D",  "Tension et courant de batterie, alimentation principale coupée, mode surveillance"),
        ("E",  "Tension et courant de batterie, alimentation principale coupée, pleine charge"),
        ("F",  "Courant de charge"),
        ("G",  "Absence de dommages matériels"),
        ("H",  "Bornes nettoyées et lubrifiées"),
        ("I",  "Bornes serrées"),
        ("J",  "Niveau d'électrolyte correct"),
        ("K",  "Densité de l'électrolyte conforme aux spécifications du fabricant"),
        ("L",  "Aucune fuite d'électrolyte"),
        ("M",  "Ventilation adéquate"),
        ("N",  "Code dateur du fabricant ou date de mise en service"),
        ("O",  "Débranchement provoque signal de défectuosité"),
        ("Q",  "Capacité de la batterie calculée"),
        ("R",  "Après la fin des essais, tension aux bornes de la batterie"),
        ("S",  "Après les essais, la tension de la batterie n'est pas inférieure à 85 % de la tension nominale"),
        ("T",  "Le générateur fournit l'alimentation au circuit C.A. qui dessert le réseau avertisseur d'incendie"),
        ("U",  "Une situation de défectuosité au générateur d'urgence provoque un signal de défectuosité sonore commun ainsi qu'une indication visuelle"),
    ],
    "e2_6": [
        ("A",  "Fonctionnement de l'indicateur de mise sous tension"),
        ("B",  "Zones d'entrée individuelles d'alarme et de surveillance indiquées clairement, de manière distincte"),
        ("C",  "Étiquettes de désignation des zones individuelles d'alarme et de surveillance correctement marquées"),
        ("D",  "Fonctionnement du signal de défectuosité commun"),
        ("E",  "Fonctionnement de l'essai d'indicateur visuel (essai de lampe)"),
        ("F",  "Surveillance du câblage d'entrée du poste de contrôle ou du répondeur"),
        ("G",  "Fonctionnement de l'indicateur visuel d'interruption du signal d'alarme sonore"),
        ("H",  "Contacts des fonctions auxiliaires fonctionnant conformément à la conception et aux spécifications"),
        ("I",  "Fonctionnement des autres indicateurs visuels des fonctions auxiliaires"),
        ("J",  "Actionnement manuel du signal d'alarme et indication"),
        ("K",  "Affichages visibles dans le lieu de l'installation"),
        ("L",  "Fonctionnement sur l'alimentation de secours"),
    ],
    "e2_7": [
        ("A",  "Fonctionnement de l'indicateur de mise sous tension"),
        ("B",  "Fonctionnement de l'indication de zone individuelle d'alarme et de surveillance"),
        ("C",  "Étiquettes de désignation des zones individuelles d'alarme et de surveillance correctement marquées"),
        ("D",  "Fonctionnement du signal de défectuosité commun"),
        ("E",  "Fonctionnement de l'essai d'indicateur visuel (essai de lampe)"),
        ("F",  "Surveillance du câblage d'entrée du poste de contrôle ou du répondeur"),
        ("G",  "Fonctionnement de l'indicateur visuel d'interruption du signal d'alarme sonore"),
        ("H",  "Contacts des fonctions auxiliaires fonctionnant conformément à la conception et aux spécifications"),
        ("I",  "Fonctionnement des autres indicateurs visuels des fonctions auxiliaires"),
        ("J",  "Actionnement manuel du signal d'alarme et indication"),
        ("K",  "Affichages visibles dans le lieu de l'installation"),
    ],
    "e2_8": [
        ("A",  "Surveillance du câblage d'entrée du poste de contrôle ou du répondeur"),
        ("B",  "Fonctionnement du signal visuel de défectuosité"),
        ("C",  "Fonctionnement du signal sonore de défectuosité"),
        ("D",  "Fonctionnement de l'interruption du signal sonore de défectuosité"),
    ],
    "e2_9": [
        ("A",  "Fonctionnement de l'imprimante selon la conception et les spécifications"),
        ("B",  "Impression correcte de la zone de chaque dispositif de déclenchement d'alarme"),
        ("C",  "Alimentation à la tension nominale"),
    ],
    "e2_10": [
        ("A",     "Confirmer la réception d'un signal de défectuosité par le poste de contrôle en cas de boucle ouverte pour chaque liaison de données"),
        ("B",     "Si des modules d'isolation en cas de défaut font partie de liaisons de données, court-circuiter le câblage et confirmer l'annonce de la défectuosité"),
        ("C_i",   "Poste de contrôle et poste de contrôle"),
        ("C_ii",  "Poste de contrôle et répondeur"),
        ("C_iii", "Répondeur et répondeur"),
    ],
}


# ── Permissions ──────────────────────────────────────────────────────────
class EstSuperviseur(permissions.BasePermission):
    def has_permission(self, request, view):
        return bool(request.user and request.user.est_superviseur())


class EstSuperviseurOuTechnicien(permissions.BasePermission):
    def has_permission(self, request, view):
        return bool(
            request.user and (request.user.est_superviseur() or request.user.est_technicien())
        )


# ── Client ───────────────────────────────────────────────────────────────
class ClientViewSet(viewsets.ModelViewSet):
    """Gestion des entreprises clientes — réservée au superviseur."""

    serializer_class = ClientSerializer
    permission_classes = [permissions.IsAuthenticated, EstSuperviseur]

    def get_queryset(self):
        return Client.objects.filter(organisation=self.request.user.organisation)

    def perform_create(self, serializer):
        serializer.save(organisation=self.request.user.organisation)


# ── Bâtiment ─────────────────────────────────────────────────────────────
class BatimentViewSet(viewsets.ModelViewSet):
    serializer_class = BatimentSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        qs = Batiment.objects.select_related("client").filter(client__organisation=user.organisation)
        if user.est_citoyen():
            qs = qs.filter(proprietaire=user)
        elif user.est_technicien():
            qs = qs.filter(rapports__techniciens=user).distinct()

        client_id = self.request.query_params.get("client")
        if client_id:
            qs = qs.filter(client_id=client_id)
        return qs

    def get_permissions(self):
        if self.action in ["create", "update", "partial_update", "destroy"]:
            return [permissions.IsAuthenticated(), EstSuperviseur()]
        return super().get_permissions()


def _html_certificat_incendie(rapport) -> str:
    """HTML du certificat d'inspection annuelle (réseau d'alarme incendie) —
    même chrome (ligne rouge, bandeau noir, pied de page bouclier) que le
    certificat extincteurs, avec son propre contenu (inventaire des
    dispositifs + conformité E1)."""
    from .pdf_design import CSS_DOCUMENT, ICONE_CALENDRIER, ICONE_PERSONNE, ICONE_PIN, entete, icone, pied_de_page

    cert = rapport.certificat
    bat = rapport.batiment
    adresse = f"{bat.numero_civique} {bat.rue}, {bat.ville}"
    if bat.code_postal:
        adresse += f"  {bat.code_postal}"
    langue = bat.client.organisation.langue
    t = lambda cle: _t(langue, cle)

    date_insp = _date_fr(rapport.date_inspection)
    date_cert = _date_fr(cert.date_emission)
    techniciens = list(rapport.techniciens.all())
    tech_noms = ", ".join(t2.get_full_name() or t2.username for t2 in techniciens) or "—"
    e1 = rapport.fiche_e1 if hasattr(rapport, "fiche_e1") else None

    dispositifs = list(rapport.dispositifs.all())
    type_counts = Counter(d.get_type_dispositif_display() or "—" for d in dispositifs)
    total = sum(type_counts.values())

    inv_rows = "".join(
        f"<tr><td>{ty}</td><td class='center bold'>{c}</td></tr>"
        for ty, c in sorted(type_counts.items())
    ) or "<tr><td colspan='2' class='muted center'>Aucun dispositif enregistré</td></tr>"

    def conf_item(condition, label):
        ok = condition and bool(condition)
        ic = "✔" if ok else "✖"
        color = "#0d6b4f" if ok else "#e11324"
        return f"<div class='conf-item'><span style='color:{color};font-weight:900;font-size:11pt;flex-shrink:0;'>{ic}</span><span>{label}</span></div>"

    conf_html = ""
    if e1:
        conf_html += conf_item(e1.inspection_essai_conforme, "Inspection et mise à l'essai conforme à la norme CAN/ULC-S536")
        conf_html += conf_item(e1.reseau_fonctionnel, "Réseau surveillé complètement fonctionnel")
        conf_html += conf_item(not e1.lacunes_constatees if e1.lacunes_constatees is not None else None, "Aucune lacune constatée sur le réseau")
        conf_html += conf_item(e1.documentation_sur_place, "Documentation du réseau présente sur place")
        if e1.commentaires:
            conf_html += f"<div style='margin-top:6px;font-size:8.5pt;color:#555;font-style:italic;'>Commentaires : {e1.commentaires}</div>"

    logo_content = organisation_logo_content(bat.client.organisation, 46)
    organisation_nom = bat.client.organisation.nom
    emetteur = cert.emis_par.get_full_name() or cert.emis_par.username if cert.emis_par else "—"
    conforme = cert.conforme
    conf_badge_bg = "#dcfce7" if conforme else "#fee2e2"
    conf_badge_color = "#16a34a" if conforme else "#e11324"
    conf_badge_texte = t("conforme_badge") if conforme else t("non_conforme_badge")

    entete_html = entete(
        logo_content, organisation_nom, f"{t('inspection_certification')} — Norme CAN/ULC-S536",
        t("certificat_no"), cert.numero, t("date_inspection"), date_insp, t("technicien_s"), tech_noms,
    )

    return f"""<!DOCTYPE html>
<html lang="{langue}">
<head>
<meta charset="UTF-8">
<title>{t("certificat_no")} {cert.numero}</title>
<style>{CSS_DOCUMENT}</style>
</head>
<body>
<div class="no-print" style="text-align:right;padding:8px 12px;background:#f8fafc;border-bottom:1px solid #e5e7eb;">
  <button onclick="window.print()" style="background:#0a0b0d;color:#fff;border:none;padding:8px 20px;border-radius:4px;font-weight:700;cursor:pointer;font-size:10pt;">{t("imprimer_pdf")}</button>
</div>
<div style="padding:20px 24px;">
{entete_html}
<div class="title-banner">
  <h2>Certificat d'inspection annuelle</h2>
  <div style="width:140px;height:1.5px;background:linear-gradient(90deg, transparent, #e11324, transparent);margin:6px auto;"></div>
  <p>Réseau d'alarme incendie — CAN/ULC-S536</p>
</div>
<div style="text-align:center;margin-bottom:14px;">
  <span style="display:inline-block;background:{conf_badge_bg};border:1.5px solid {conf_badge_color};color:{conf_badge_color};font-size:11pt;font-weight:900;letter-spacing:2px;padding:5px 22px;border-radius:100px;">{conf_badge_texte}</span>
  {'<p style="margin-top:6px;font-size:8pt;color:#e11324;">Des réparations sont requises avant que ce certificat ne soit conforme.</p>' if not conforme else ''}
</div>
<div style="text-align:left;margin-bottom:6px;">
  <div class="card-title">{t("client")}</div>
  <div class="card-main" style="font-size:11pt;">{bat.client.nom}</div>
</div>
<div style="text-align:center;margin-bottom:6px;">
  <div class="card-title">{t("adresse_inspectee")}</div>
</div>
<div class="info-card" style="display:flex;align-items:center;justify-content:center;gap:14px;margin-bottom:18px;">
  <span style="display:inline-flex;align-items:center;justify-content:center;width:36px;height:36px;border-radius:50%;background:#f1f5f9;flex-shrink:0;">{icone(ICONE_PIN, 16, '#6b7280')}</span>
  <div class="card-main" style="font-size:20pt; font-weight:900;">{adresse}</div>
</div>
<div class="sec-title">Inventaire des dispositifs</div>
<table>
  <thead><tr><th>Type de dispositif</th><th class="center">Qté</th></tr></thead>
  <tbody>{inv_rows}<tr style="font-weight:700;background:#f8fafc;border-top:1.5px solid #e5e7eb;"><td>Total</td><td class="center bold">{total}</td></tr></tbody>
</table>
<div class="sec-title">Conformité — Mise à l'essai</div>
<div class="conf-box">{conf_html or '<p style="color:#9ca3af;font-style:italic;font-size:9pt;">Données E1 non disponibles.</p>'}</div>
<div class="sig-row">
  <div class="sig-block" style="display:flex;align-items:center;gap:10px;">
    <span class="sig-icon">{icone(ICONE_PERSONNE, 14, '#e11324')}</span>
    <div>
      <div class="sig-label">Superviseur / Responsable</div>
      <div class="sig-name">{emetteur}</div>
      <div style="font-size:8pt;color:#555;">{organisation_nom}</div>
    </div>
  </div>
  <div class="sig-block" style="display:flex;align-items:center;gap:10px;">
    <span class="sig-icon">{icone(ICONE_CALENDRIER, 14, '#e11324')}</span>
    <div>
      <div class="sig-label">Date d'émission</div>
      <div class="sig-name">{date_cert}</div>
      <div style="font-size:8pt;color:#555;">{t("certificat_no")} {cert.numero}</div>
    </div>
  </div>
</div>
{pied_de_page(organisation_nom, t("footer_certificat_incendie"))}
</div>
</body>
</html>"""


def _html_rapport_incendie_complet(rapport) -> str:
    """HTML du rapport technique complet (E1 + E2 + légende + E3) — même
    chrome (ligne rouge, bandeau noir, pied de page bouclier) que le
    certificat, avec le contenu technique détaillé."""
    from .pdf_design import CSS_DOCUMENT, entete, pied_de_page

    bat = rapport.batiment
    adresse = f"{bat.numero_civique} {bat.rue}, {bat.ville}"
    fabricant_panneau = bat.fabricant_reseau or "—"
    modele_panneau = bat.modele_systeme or "—"
    langue = bat.client.organisation.langue
    t = lambda cle: _t(langue, cle)
    date_insp = _date_fr(rapport.date_inspection)
    techniciens = list(rapport.techniciens.all())
    tech_noms = ", ".join(t2.get_full_name() or t2.username for t2 in techniciens) or "—"

    e1 = rapport.fiche_e1 if hasattr(rapport, "fiche_e1") else None
    e2 = rapport.fiche_e2 if hasattr(rapport, "fiche_e2") else None
    legende = rapport.fiche_legende.dispositifs if hasattr(rapport, "fiche_legende") else {}
    dispositifs = list(rapport.dispositifs.select_related("section").all())

    type_counts = Counter(d.get_type_dispositif_display() or "—" for d in dispositifs)
    total_disp = sum(type_counts.values())
    inv_rows = "".join(
        f"<tr><td>{ty}</td><td class='center bold'>{c}</td></tr>"
        for ty, c in sorted(type_counts.items())
    ) or "<tr><td colspan='2' class='muted center'>Aucun dispositif</td></tr>"

    e1_html = ""
    if e1:
        CHAMPS_E1 = [
            ("A", "Fonctionnement en une étape",                        e1.fonctionnement_une_etape),
            ("B", "Fonctionnement en deux étapes",                      e1.fonctionnement_deux_etapes),
            ("C", "Inspection et mise à l'essai (CAN/ULC-S536)",       e1.inspection_essai_conforme),
            ("D", "Documentation du réseau sur place",                  e1.documentation_sur_place),
            ("E", "Réseau fonctionnel",                                 e1.reseau_fonctionnel),
            ("F", "Lacunes constatées",                                 e1.lacunes_constatees),
            ("H", "Copie remise au responsable",                        e1.copie_remise_responsable),
        ]
        rows = "".join(
            f"<tr><td class='bold' style='width:30px;'>{l}</td><td>{label}</td><td class='center'>{_val_oui_non(v)}</td></tr>"
            for l, label, v in CHAMPS_E1
        )
        e1_html = f"""<div class="sec-title">E1 — Rapport annuel de mise à l'essai</div>
<table><thead><tr><th></th><th>Champ</th><th class='center'>Valeur</th></tr></thead><tbody>{rows}</tbody></table>
{'<p style="margin-top:6px;font-size:9pt;"><strong>Commentaires :</strong> ' + (e1.commentaires or '—') + '</p>' if e1 else ''}"""

    e2_html = ""
    e2_parts = []
    for key, titre in E2_TITRES.items():
        sec_data = (e2.details or {}).get(key, {}) if (e2 and e2.details) else {}
        items_defs = E2_ITEMS.get(key, [])
        loc = sec_data.get("localisation", "")
        loc_str = f" <span style='font-size:8pt;color:#444;'>({loc})</span>" if loc else ""
        if key in ("e2_11", "e2_12"):
            val = sec_data.get("remarques", "")
            content = f"<p style='font-size:9pt;color:#111;padding:4px 0;'>{val or '—'}</p>"
        else:
            rows = "".join(
                f"<tr>"
                f"<td class='bold' style='width:36px;vertical-align:top;'>{iid}</td>"
                f"<td style='line-height:1.4;'>{lbl}</td>"
                f"<td class='center' style='width:64px;vertical-align:top;'>{_val_so(sec_data.get(iid))}</td>"
                f"</tr>"
                for iid, lbl in items_defs
            )
            content = (
                f"<table><thead><tr>"
                f"<th style='width:36px;'></th>"
                f"<th>Élément vérifié</th>"
                f"<th class='center' style='width:64px;'>Résultat</th>"
                f"</tr></thead><tbody>{rows}</tbody></table>"
            )
        e2_parts.append(
            f"<div style='margin-bottom:14px;page-break-inside:avoid;'>"
            f"<div class='sec-sub'>{titre}{loc_str}</div>"
            f"{content}"
            f"</div>"
        )
    if e2_parts:
        e2_html = '<div class="sec-title">E2 — Essai du poste de contrôle</div>' + "".join(e2_parts)

    legende_rows = "".join(
        f"<tr><td class='bold'>{code}</td><td>{desc}</td>"
        f"<td>{(legende.get(code) or {}).get('type') or '—'}</td>"
        f"<td>{(legende.get(code) or {}).get('modele') or '—'}</td></tr>"
        for code, desc in LEGENDE_DISPOSITIFS
    )
    legende_html = (
        '<div class="sec-title">Légende des dispositifs</div>'
        "<table><thead><tr><th>Dispositif</th><th>Description</th><th>Type</th><th>No de modèle</th></tr></thead>"
        f"<tbody>{legende_rows}</tbody></table>"
    )

    sections_html = ""
    for section in rapport.sections.prefetch_related("dispositifs").all():
        devs = list(section.dispositifs.all())
        if not devs:
            continue
        rows = ""
        for d in devs:
            is_defect = d.est_defectueux
            is_ni = not is_defect and d.annonce_statut == "NI"
            bg = ' style="background:#fef2f2;"' if is_defect else ' style="background:#fef3c7;"' if is_ni else ""
            a = "1" if d.installation_correcte is True else "S.O." if d.installation_correcte is None else "0"
            b = "1" if d.necessite_entretien is True else "S.O." if d.necessite_entretien is None else "0"
            c_val = "1" if d.alarme_confirmee is True else "S.O." if d.alarme_confirmee is None else "0"
            d_val = d.annonce_statut or "—"
            e_val = d.zone_circuit or "—"
            loc_style = ' style="color:#cc0000;font-weight:700;"' if is_defect else ' style="color:#b45309;font-weight:700;"' if is_ni else ""
            rows += (
                f"<tr{bg}>"
                f"<td{loc_style}>{d.localisation}</td>"
                f"<td class='center bold'>{d.type_dispositif or '—'}</td>"
                f"<td class='center'>{a}</td>"
                f"<td class='center'>{b}</td>"
                f"<td class='center'>{c_val}</td>"
                f"<td class='center'>{d_val}</td>"
                f"<td class='center'>{e_val}</td>"
                f"<td>{d.remarque or ''}</td>"
                f"</tr>"
            )
        sections_html += (
            f"<div class='sec-sub'>{section.nom}</div>"
            f"<table>"
            f"<thead><tr>"
            f"<th>Localisation</th>"
            f"<th class='center'>Type</th>"
            f"<th class='center' title='Installation correcte'>A</th>"
            f"<th class='center' title='Nécessite entretien'>B</th>"
            f"<th class='center' title='Alarme confirmée'>C</th>"
            f"<th class='center' title='D=Défectueux, I=Inspecté, NI=Non inspecté'>D</th>"
            f"<th class='center' title='Zone / Circuit'>E</th>"
            f"<th>Remarque</th>"
            f"</tr></thead>"
            f"<tbody>{rows}</tbody></table>"
        )
    if not sections_html:
        sections_html = "<p class='muted' style='font-size:9pt;'>Aucun dispositif enregistré.</p>"

    cert_html = ""
    if hasattr(rapport, "certificat"):
        c = rapport.certificat
        badge_color = "#0d6b4f" if c.conforme else "#e11324"
        badge_bg = "#e9f6f2" if c.conforme else "#fef2f2"
        badge_texte = "Conforme" if c.conforme else "Non conforme"
        cert_html = (
            f'<div style="display:flex;align-items:center;gap:8px;background:#fff7ed;border:1.5px solid #ff6b1a;'
            f'border-radius:6px;padding:8px 14px;margin-top:12px;">'
            f'<span style="font-size:7pt;font-weight:700;text-transform:uppercase;color:#9a4a13;">Certificat</span>'
            f'<span style="font-size:11pt;font-weight:900;color:#ff6b1a;">{c.numero}</span>'
            f'<span style="font-size:8pt;color:#555;">· Émis le {_date_fr(c.date_emission)}</span>'
            f'<span style="background:{badge_bg};color:{badge_color};font-size:7pt;font-weight:700;padding:2px 8px;border-radius:100px;border:1px solid {badge_color};">{badge_texte}</span>'
            f'{"<span style=\"background:#0d6b4f;color:#fff;font-size:7pt;font-weight:700;padding:2px 7px;border-radius:100px;\">Envoyé</span>" if c.certificat_envoye else ""}'
            f'</div>'
        )

    logo_content = organisation_logo_content(bat.client.organisation, 46)
    organisation_nom = bat.client.organisation.nom

    entete_html = entete(
        logo_content, organisation_nom, "Rapport d'inspection annuelle — CAN/ULC-S536",
        t("certificat_no"), (rapport.certificat.numero if hasattr(rapport, "certificat") else "—"),
        t("date_inspection"), date_insp, t("technicien_s"), tech_noms,
    )

    return f"""<!DOCTYPE html>
<html lang="{langue}">
<head>
<meta charset="UTF-8">
<title>Rapport d'inspection — {adresse}</title>
<style>{CSS_DOCUMENT}
  .info-grid{{ display:grid; grid-template-columns:1fr 1fr 1fr; gap:10px; margin-bottom:18px; }}
</style>
</head>
<body>
<div class="no-print" style="text-align:right;padding:8px 12px;background:#f8fafc;border-bottom:1px solid #e5e7eb;">
  <button onclick="window.print()" style="background:#0a0b0d;color:#fff;border:none;padding:8px 20px;border-radius:4px;font-weight:700;cursor:pointer;font-size:10pt;">{t("imprimer_pdf")}</button>
</div>
<div style="padding:20px 24px;">
{entete_html}
<div class="title-banner">
  <h2>Rapport d'inspection annuelle</h2>
  <div style="width:140px;height:1.5px;background:linear-gradient(90deg, transparent, #e11324, transparent);margin:6px auto;"></div>
  <p>Réseau d'alarme incendie — CAN/ULC-S536</p>
</div>
<div class="info-grid">
  <div class="info-card">
    <div class="card-title">Adresse</div>
    <div class="card-main">{adresse}</div>
  </div>
  <div class="info-card">
    <div class="card-title">Fabricant du panneau</div>
    <div class="card-main">{fabricant_panneau}</div>
  </div>
  <div class="info-card">
    <div class="card-title">Modèle du panneau</div>
    <div class="card-main">{modele_panneau}</div>
  </div>
</div>
{e1_html}
{e2_html}
{legende_html}
<div class="sec-title">E3 — Inventaire global</div>
<table>
  <thead><tr><th>Type de dispositif</th><th class="center">Qté</th></tr></thead>
  <tbody>{inv_rows}<tr style="font-weight:700;background:#f8fafc;border-top:1.5px solid #e5e7eb;"><td>Total</td><td class="center bold">{total_disp}</td></tr></tbody>
</table>
<div class="sec-title">E3 — Détail par section</div>
<div style="font-size:7.5pt;color:#000;margin-bottom:8px;font-style:italic;">A = Installation correcte &nbsp;|&nbsp; B = Nécessite entretien &nbsp;|&nbsp; C = Alarme confirmée &nbsp;|&nbsp; D = Statut (D=Défectueux, I=Inspecté, NI=Non inspecté) &nbsp;|&nbsp; E = Zone/Circuit &nbsp;&nbsp;(A/B/C : 1 = Oui, 0 = Non)</div>
{sections_html}
{cert_html}
{pied_de_page(organisation_nom, t("footer_rapport_incendie"))}
</div>
</body>
</html>"""


# ── Rapport ──────────────────────────────────────────────────────────────
class EstModuleRapportIncendieActif(permissions.BasePermission):
    message = "Le module « Rapport d'inspection incendie » n'est pas activé pour votre organisation."

    def has_permission(self, request, view):
        organisation = getattr(request.user, "organisation", None)
        return bool(organisation and organisation.a_le_module("rapport_incendie"))


class RapportViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated, EstModuleRapportIncendieActif]

    def get_serializer_class(self):
        if self.action == "list":
            return RapportListSerializer
        if self.action == "create":
            return RapportCreateSerializer
        return RapportDetailSerializer

    def get_queryset(self):
        user = self.request.user
        qs = Rapport.objects.select_related("batiment", "batiment__client", "cree_par", "citoyen").prefetch_related("techniciens").filter(batiment__client__organisation=user.organisation)

        if user.est_citoyen():
            qs = qs.filter(citoyen=user)
        elif user.est_technicien():
            qs = qs.filter(techniciens=user)
        # le superviseur voit tout

        # Filtres — client, direction, type d'application, statut
        client_id = self.request.query_params.get("client")
        direction = self.request.query_params.get("direction")
        application = self.request.query_params.get("application")
        statut = self.request.query_params.get("statut")
        if client_id:
            qs = qs.filter(batiment__client_id=client_id)
        if direction:
            qs = qs.filter(batiment__direction__icontains=direction)
        if application:
            qs = qs.filter(batiment__type_application=application)
        if statut:
            qs = qs.filter(statut=statut)

        return qs.distinct()

    def get_permissions(self):
        if self.action in ["create", "destroy", "reassigner", "rouvrir"]:
            return [permissions.IsAuthenticated(), EstSuperviseur()]
        if self.action in ["update", "partial_update"]:
            return [permissions.IsAuthenticated(), EstSuperviseurOuTechnicien()]
        return super().get_permissions()

    @action(detail=True, methods=["patch"])
    def reassigner(self, request, pk=None):
        """Superviseur seulement — change le bâtiment et/ou les techniciens assignés après création."""
        rapport = self.get_object()
        changements = []

        if "batiment" in request.data:
            try:
                nouveau_batiment = Batiment.objects.get(pk=request.data["batiment"])
            except (Batiment.DoesNotExist, TypeError, ValueError):
                return Response({"error": "Bâtiment introuvable."}, status=status.HTTP_400_BAD_REQUEST)
            if nouveau_batiment.id != rapport.batiment_id:
                rapport.batiment = nouveau_batiment
                changements.append(f"Bâtiment changé pour {nouveau_batiment.adresse_complete}")

        if "techniciens" in request.data:
            rapport.techniciens.set(request.data.get("techniciens") or [])
            changements.append("Techniciens réassignés")

        if "citoyen" in request.data:
            citoyen_id = request.data.get("citoyen")
            if citoyen_id:
                try:
                    nouveau_citoyen = Utilisateur.objects.get(pk=citoyen_id, role=Utilisateur.Role.CITOYEN)
                except (Utilisateur.DoesNotExist, TypeError, ValueError):
                    return Response({"error": "Citoyen introuvable."}, status=status.HTTP_400_BAD_REQUEST)
                rapport.citoyen = nouveau_citoyen
                changements.append(f"Citoyen réassigné à {nouveau_citoyen.username}")
            else:
                rapport.citoyen = None
                changements.append("Citoyen retiré du rapport")

        if changements:
            rapport.save()
            for c in changements:
                rapport.historiser(request.user, c)

        return Response(RapportDetailSerializer(rapport).data)

    @action(detail=True, methods=["post"])
    def rouvrir(self, request, pk=None):
        rapport = self.get_object()
        if not request.user.est_superviseur():
            return Response(
                {"error": "Seul le superviseur peut rouvrir un rapport."},
                status=status.HTTP_403_FORBIDDEN,
            )
        if rapport.statut != Rapport.Statut.FERME:
            return Response({"error": "Ce rapport est déjà ouvert."}, status=status.HTTP_400_BAD_REQUEST)

        rapport.rouvrir(request.user)
        return Response(RapportDetailSerializer(rapport).data)

    def perform_create(self, serializer):
        rapport = serializer.save(cree_par=self.request.user)
        # Crée automatiquement les fiches E1/E2/légende vides, prêtes à être remplies
        FicheE1.objects.create(rapport=rapport)
        FicheE2.objects.create(rapport=rapport)
        FicheLegende.objects.create(rapport=rapport)
        rapport.historiser(self.request.user, "Rapport créé")

        # Crée automatiquement le rapport de vérification des extincteurs
        # portatifs correspondant — même adresse, même citoyen, mêmes
        # techniciens assignés au départ (le superviseur peut ensuite les
        # réassigner indépendamment sur ce rapport).
        rapport_extincteur = RapportExtincteur.objects.create(
            batiment=rapport.batiment,
            rapport_alarme=rapport,
            cree_par=self.request.user,
            citoyen=rapport.citoyen,
            date_inspection=rapport.date_inspection,
        )
        rapport_extincteur.techniciens.set(rapport.techniciens.all())
        rapport_extincteur.historiser(self.request.user, "Rapport créé automatiquement avec le rapport principal")

        _creer_rapport_eclairage_lie(rapport_extincteur, self.request.user)

    def perform_update(self, serializer):
        instance = self.get_object()
        if instance.statut == Rapport.Statut.FERME and not self.request.user.est_superviseur():
            from rest_framework.exceptions import ValidationError
            raise ValidationError("Ce rapport est fermé et ne peut plus être modifié.")
        rapport = serializer.save()
        rapport.historiser(self.request.user, "Rapport modifié")

    @action(detail=True, methods=["post"])
    def fermer(self, request, pk=None):
        rapport = self.get_object()
        if not request.user.est_superviseur():
            return Response(
                {"error": "Seul le superviseur peut fermer un rapport."},
                status=status.HTTP_403_FORBIDDEN,
            )
        if rapport.statut == Rapport.Statut.FERME:
            return Response({"error": "Ce rapport est déjà fermé."}, status=status.HTTP_400_BAD_REQUEST)

        rapport.fermer(request.user)
        return Response(RapportDetailSerializer(rapport).data)

    @action(detail=True, methods=["post"], url_path="envoyer-certificat")
    def envoyer_certificat(self, request, pk=None):
        rapport = self.get_object()
        if not request.user.est_superviseur():
            return Response(
                {"error": "Seul le superviseur peut envoyer le certificat."},
                status=status.HTTP_403_FORBIDDEN,
            )
        if rapport.statut != Rapport.Statut.FERME:
            return Response(
                {"error": "Le rapport doit être fermé avant d'envoyer le certificat."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if not hasattr(rapport, "certificat"):
            return Response({"error": "Aucun certificat trouvé pour ce rapport."}, status=status.HTTP_404_NOT_FOUND)

        client = rapport.batiment.client
        if client.mode_livraison == Client.ModeLivraison.DIRECT:
            from .emailing import envoyer_certificats_directs_batiment

            ok, message = envoyer_certificats_directs_batiment(rapport.batiment, request.user)
            if not ok:
                return Response({"error": message}, status=status.HTTP_400_BAD_REQUEST)
            return Response({"message": message})

        rapport.certificat.certificat_envoye = True
        rapport.certificat.save()
        rapport.historiser(request.user, f"Certificat envoyé au citoyen {rapport.citoyen.username if rapport.citoyen else '—'}")

        if rapport.citoyen and rapport.citoyen.email:
            from .emailing import envoyer_email_certificat_disponible

            envoyer_email_certificat_disponible(rapport)

        return Response({"message": "Certificat envoyé au citoyen."})

    @action(detail=True, methods=["get", "patch"], url_path="fiche-e1")
    def fiche_e1(self, request, pk=None):
        rapport = self.get_object()
        if request.method == "GET":
            return Response(FicheE1Serializer(rapport.fiche_e1).data)

        if rapport.statut == Rapport.Statut.FERME and not request.user.est_superviseur():
            return Response({"error": "Ce rapport est fermé."}, status=status.HTTP_400_BAD_REQUEST)
        serializer = FicheE1Serializer(rapport.fiche_e1, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        rapport.historiser(request.user, "Fiche E1 modifiée")
        return Response(serializer.data)

    @action(detail=True, methods=["get", "patch"], url_path="fiche-e2")
    def fiche_e2(self, request, pk=None):
        rapport = self.get_object()
        if request.method == "GET":
            return Response(FicheE2Serializer(rapport.fiche_e2).data)

        if rapport.statut == Rapport.Statut.FERME and not request.user.est_superviseur():
            return Response({"error": "Ce rapport est fermé."}, status=status.HTTP_400_BAD_REQUEST)
        serializer = FicheE2Serializer(rapport.fiche_e2, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        rapport.historiser(request.user, "Fiche E2 modifiée")
        return Response(serializer.data)

    @action(detail=True, methods=["get", "patch"], url_path="fiche-legende")
    def fiche_legende(self, request, pk=None):
        rapport = self.get_object()
        fiche, _ = FicheLegende.objects.get_or_create(rapport=rapport)

        if request.method == "GET":
            return Response(FicheLegendeSerializer(fiche).data)

        if rapport.statut == Rapport.Statut.FERME and not request.user.est_superviseur():
            return Response({"error": "Ce rapport est fermé."}, status=status.HTTP_400_BAD_REQUEST)
        serializer = FicheLegendeSerializer(fiche, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    @action(detail=True, methods=["get", "post"])
    def sections(self, request, pk=None):
        """Le technicien crée les sections dont il a besoin (ex. un étage à la fois)."""
        rapport = self.get_object()

        if request.method == "GET":
            return Response(SectionDispositifSerializer(rapport.sections.all(), many=True).data)

        if rapport.statut == Rapport.Statut.FERME and not request.user.est_superviseur():
            return Response({"error": "Ce rapport est fermé."}, status=status.HTTP_400_BAD_REQUEST)

        serializer = SectionDispositifSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save(rapport=rapport)
        rapport.historiser(request.user, f"Section ajoutée : {serializer.validated_data.get('nom')}")
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=["get", "post"])
    def dispositifs(self, request, pk=None):
        rapport = self.get_object()

        if request.method == "GET":
            return Response(DispositifSerializer(rapport.dispositifs.all(), many=True).data)

        if rapport.statut == Rapport.Statut.FERME and not request.user.est_superviseur():
            return Response(
                {"error": "Ce rapport est fermé, impossible d'ajouter un dispositif."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        serializer = DispositifSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        # La section (si fournie) doit appartenir à ce même rapport
        section = serializer.validated_data.get("section")
        if section and section.rapport_id != rapport.id:
            return Response({"error": "Cette section n'appartient pas à ce rapport."}, status=status.HTTP_400_BAD_REQUEST)

        serializer.save(rapport=rapport)
        rapport.historiser(request.user, f"Dispositif ajouté : {serializer.validated_data.get('localisation')}")
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=["get"])
    def historique(self, request, pk=None):
        rapport = self.get_object()
        return Response(HistoriqueRapportSerializer(rapport.historique.all(), many=True).data)

    @action(detail=False, methods=["get"])
    def aujourdhui(self, request):
        """Rapports assignés au technicien connecté, pour la date du jour."""
        qs = self.get_queryset().filter(date_inspection=date.today())
        return Response(RapportListSerializer(qs, many=True).data)

    @action(detail=False, methods=["get"])
    def stats(self, request):
        qs = self.get_queryset()
        return Response({
            "total": qs.count(),
            "ouverts": qs.filter(statut=Rapport.Statut.OUVERT).count(),
            "fermes": qs.filter(statut=Rapport.Statut.FERME).count(),
            "lacunes_ouvertes": Dispositif.objects.filter(
                rapport__in=qs, necessite_entretien=True
            ).count(),
        })

    @action(detail=True, methods=["get"], url_path="certificat-pdf")
    def certificat_pdf(self, request, pk=None):
        rapport = self.get_object()
        if rapport.statut != Rapport.Statut.FERME:
            return Response({"error": "Le rapport doit être fermé."}, status=status.HTTP_400_BAD_REQUEST)
        if not hasattr(rapport, "certificat"):
            return Response({"error": "Aucun certificat pour ce rapport."}, status=status.HTTP_404_NOT_FOUND)

        html = _html_certificat_incendie(rapport)
        return HttpResponse(html, content_type="text/html; charset=utf-8")

    @action(detail=True, methods=["get"], url_path="telecharger")
    def telecharger(self, request, pk=None):
        rapport = self.get_object()
        html = _html_rapport_incendie_complet(rapport)
        return HttpResponse(html, content_type="text/html; charset=utf-8")

    @action(detail=True, methods=["get"])
    def excel(self, request, pk=None):
        rapport = self.get_object()
        bat = rapport.batiment
        adresse = f"{bat.numero_civique} {bat.rue}, {bat.ville}"
        techniciens = ", ".join(
            t.get_full_name() or t.username for t in rapport.techniciens.all()
        ) or "—"
        e1 = rapport.fiche_e1 if hasattr(rapport, "fiche_e1") else None
        e2 = rapport.fiche_e2 if hasattr(rapport, "fiche_e2") else None
        legende = rapport.fiche_legende.dispositifs if hasattr(rapport, "fiche_legende") else {}
        dispositifs = list(rapport.dispositifs.select_related("section").all())
        type_counts = Counter(d.get_type_dispositif_display() or "—" for d in dispositifs)

        def oui_non(v):
            return "Oui" if v is True else "Non" if v is False else "—"

        def so_valeur(v):
            if v == "oui": return "Oui"
            if v == "non": return "Non"
            if v == "sans_objet": return "S.O."
            return v or "—"

        wb = excel_workbook()
        ws = wb.active
        ws.title = "Rapport incendie"
        # Largeurs fixées d'avance : le document empile plusieurs tableaux à
        # nombre de colonnes différent (E1, E2, légende, E3...), donc on ne
        # peut pas ajuster automatiquement à la fin sur un seul tableau.
        for col, largeur in zip("ABCDEFGHI", [40, 60, 20, 20, 20, 20, 16, 16, 30]):
            ws.column_dimensions[col].width = largeur
        excel_entete_rapport(
            ws, organisation_nom=bat.client.organisation.nom, adresse=adresse,
            date_insp=_date_fr(rapport.date_inspection),
            statut=rapport.get_statut_display(), techniciens=techniciens,
        )

        ligne = 7

        # ── E1 — Rapport annuel de mise à l'essai ──────────────────────────
        if e1:
            champs_e1 = [
                ("Fonctionnement en une étape", e1.fonctionnement_une_etape),
                ("Fonctionnement en deux étapes", e1.fonctionnement_deux_etapes),
                ("Inspection et mise à l'essai (CAN/ULC-S536)", e1.inspection_essai_conforme),
                ("Documentation du réseau sur place", e1.documentation_sur_place),
                ("Réseau fonctionnel", e1.reseau_fonctionnel),
                ("Lacunes constatées", e1.lacunes_constatees),
                ("Copie remise au responsable", e1.copie_remise_responsable),
            ]
            ws.cell(row=ligne, column=1, value="E1 — Rapport annuel de mise à l'essai").font = Font(bold=True, color=NAVY, size=12)
            ligne += 1
            excel_ligne_entetes(ws, ["Champ", "Valeur"], ligne=ligne)
            for label, val in champs_e1:
                ligne += 1
                ws.cell(row=ligne, column=1, value=label)
                ws.cell(row=ligne, column=2, value=oui_non(val))
            ligne += 1
            ws.cell(row=ligne, column=1, value="Commentaires")
            ws.cell(row=ligne, column=2, value=e1.commentaires or "—")
            ligne += 2

        # ── E2 — Essai du poste de contrôle ─────────────────────────────────
        ws.cell(row=ligne, column=1, value="E2 — Essai du poste de contrôle").font = Font(bold=True, color=NAVY, size=12)
        ligne += 1
        for key, titre in E2_TITRES.items():
            sec_data = (e2.details or {}).get(key, {}) if (e2 and e2.details) else {}
            loc = sec_data.get("localisation", "")
            sous_titre = f"{titre} ({loc})" if loc else titre
            ws.cell(row=ligne, column=1, value=sous_titre).font = Font(bold=True, color=RED, size=10.5)
            ligne += 1
            if key in ("e2_11", "e2_12"):
                ws.cell(row=ligne, column=1, value="Remarques")
                ws.cell(row=ligne, column=2, value=sec_data.get("remarques") or "—")
                ligne += 2
                continue
            items_defs = E2_ITEMS.get(key, [])
            excel_ligne_entetes(ws, ["", "Élément vérifié", "Résultat"], ligne=ligne)
            for iid, lbl in items_defs:
                ligne += 1
                ws.cell(row=ligne, column=1, value=iid)
                ws.cell(row=ligne, column=2, value=lbl)
                ws.cell(row=ligne, column=3, value=so_valeur(sec_data.get(iid)))
            ligne += 2

        # ── Légende des dispositifs ──────────────────────────────────────────
        ws.cell(row=ligne, column=1, value="Légende des dispositifs").font = Font(bold=True, color=NAVY, size=12)
        ligne += 1
        excel_ligne_entetes(ws, ["Dispositif", "Description", "Type", "N° de modèle"], ligne=ligne)
        for code, desc in LEGENDE_DISPOSITIFS:
            ligne += 1
            infos = legende.get(code) or {}
            ws.cell(row=ligne, column=1, value=code)
            ws.cell(row=ligne, column=2, value=desc)
            ws.cell(row=ligne, column=3, value=infos.get("type") or "—")
            ws.cell(row=ligne, column=4, value=infos.get("modele") or "—")
        ligne += 2

        # ── E3 — Inventaire global ───────────────────────────────────────────
        ws.cell(row=ligne, column=1, value="E3 — Inventaire global").font = Font(bold=True, color=NAVY, size=12)
        ligne += 1
        excel_ligne_entetes(ws, ["Type de dispositif", "Quantité"], ligne=ligne)
        for t, c in sorted(type_counts.items()):
            ligne += 1
            ws.cell(row=ligne, column=1, value=t)
            ws.cell(row=ligne, column=2, value=c)
        ligne += 1
        ws.cell(row=ligne, column=1, value="Total").font = Font(bold=True)
        ws.cell(row=ligne, column=2, value=sum(type_counts.values())).font = Font(bold=True)
        ligne += 2

        # ── E3 — Détail par section ──────────────────────────────────────────
        ws.cell(row=ligne, column=1, value="E3 — Détail par section").font = Font(bold=True, color=NAVY, size=12)
        ligne += 1
        colonnes_disp = [
            "Section", "Localisation", "Type", "Installation correcte",
            "Nécessite entretien", "Alarme confirmée", "Statut", "Zone/Circuit", "Remarque",
        ]
        excel_ligne_entetes(ws, colonnes_disp, ligne=ligne)
        for section in rapport.sections.prefetch_related("dispositifs").all():
            for d in section.dispositifs.all():
                ligne += 1
                valeurs = [
                    section.nom, d.localisation, d.type_dispositif,
                    oui_non(d.installation_correcte), oui_non(d.necessite_entretien), oui_non(d.alarme_confirmee),
                    d.annonce_statut, d.zone_circuit, d.remarque,
                ]
                for col, valeur in enumerate(valeurs, start=1):
                    ws.cell(row=ligne, column=col, value=valeur)

        ws.freeze_panes = "A7"
        return excel_reponse(wb, excel_nom_fichier("Incendie", adresse))


# ── Section (regroupement de dispositifs) ───────────────────────────────
class SectionDispositifViewSet(viewsets.ModelViewSet):
    """Accès direct à une section — pour la renommer ou la supprimer."""

    serializer_class = SectionDispositifSerializer
    permission_classes = [permissions.IsAuthenticated, EstSuperviseurOuTechnicien, EstModuleRapportIncendieActif]

    def get_queryset(self):
        user = self.request.user
        qs = SectionDispositif.objects.select_related("rapport").filter(rapport__batiment__client__organisation=user.organisation)
        if user.est_technicien():
            qs = qs.filter(rapport__techniciens=user)
        return qs.distinct()

    def perform_update(self, serializer):
        section = self.get_object()
        if section.rapport.statut == Rapport.Statut.FERME and not self.request.user.est_superviseur():
            from rest_framework.exceptions import ValidationError
            raise ValidationError("Le rapport associé est fermé.")
        serializer.save()


# ── Dispositif ───────────────────────────────────────────────────────────
class DispositifViewSet(viewsets.ModelViewSet):
    """Accès direct à un dispositif — surtout pour corriger une ligne."""

    serializer_class = DispositifSerializer
    permission_classes = [permissions.IsAuthenticated, EstSuperviseurOuTechnicien, EstModuleRapportIncendieActif]

    def get_queryset(self):
        user = self.request.user
        qs = Dispositif.objects.select_related("rapport").filter(rapport__batiment__client__organisation=user.organisation)
        if user.est_technicien():
            qs = qs.filter(rapport__techniciens=user)
        return qs.distinct()

    def perform_update(self, serializer):
        dispositif = self.get_object()
        if dispositif.rapport.statut == Rapport.Statut.FERME and not self.request.user.est_superviseur():
            from rest_framework.exceptions import ValidationError
            raise ValidationError("Le rapport associé est fermé.")
        serializer.save()


def _html_certificat_extincteur(rapport) -> str:
    """HTML du certificat de vérification extincteurs portatifs — couvre
    l'éclairage d'urgence lié le cas échéant (un seul certificat pour les
    deux équipements)."""
    from .pdf_design import CSS_DOCUMENT, ICONE_BOUCLIER, ICONE_CALENDRIER, ICONE_CUISINE, ICONE_EXTINCTEUR, ICONE_PERSONNE, ICONE_PIN, ICONE_SORTIE, case, entete, icone, icone_badge, pied_de_page

    cert = rapport.certificat
    bat = rapport.batiment
    adresse = f"{bat.numero_civique} {bat.rue}, {bat.ville}"
    if bat.code_postal:
        adresse += f"  {bat.code_postal}"

    langue = bat.client.organisation.langue
    t = lambda cle: _t(langue, cle)

    date_insp = _date_fr(rapport.date_inspection)
    date_cert = _date_fr(cert.date_emission)
    techniciens = list(rapport.techniciens.all())
    items = list(rapport.extincteurs.all())
    tech_noms = ", ".join(t2.get_full_name() or t2.username for t2 in techniciens) or "—"

    # ── Certificat unifié : une visite couvre extincteurs + éclairage
    # d'urgence en même temps (voir rapport_eclairage_lie) — un seul
    # certificat reflète donc l'état des deux équipements. Non conforme
    # dès qu'un extincteur OU une unité d'éclairage est défectueux.
    rapport_eclairage = getattr(rapport, "rapport_eclairage_lie", None)
    eclairages = list(rapport_eclairage.eclairages_urgence.all()) if rapport_eclairage else []
    est_conforme = not any(it.etat == "D" for it in items) and not any(it.etat == "D" for it in eclairages)
    conformite_bg = "#dcfce7" if est_conforme else "#fee2e2"
    conformite_color = "#16a34a" if est_conforme else "#e11324"

    def _badge_equipement(conforme, non_conforme, so):
        if so:
            return f"<span style='display:inline-block;font-size:7.5pt;font-weight:800;letter-spacing:0.5px;color:#9ca3af;background:#f3f4f6;border:1px solid #e5e7eb;border-radius:100px;padding:3px 10px;'>{t('so')}</span>"
        if non_conforme:
            return f"<span style='display:inline-block;font-size:7.5pt;font-weight:800;letter-spacing:0.5px;color:#e11324;background:#fee2e2;border:1px solid #fecaca;border-radius:100px;padding:3px 10px;'>{t('non_conforme_badge')}</span>"
        if conforme:
            return f"<span style='display:inline-block;font-size:7.5pt;font-weight:800;letter-spacing:0.5px;color:#16a34a;background:#dcfce7;border:1px solid #bbf7d0;border-radius:100px;padding:3px 10px;'>{t('conforme_badge')}</span>"
        return "<span class='muted' style='font-size:8pt;'>—</span>"

    def _ligne_equipement(nom, icone_svg, applicable, items_liste):
        so = not applicable or not items_liste
        defectueux = applicable and any(it.etat == "D" for it in items_liste)
        conforme = applicable and bool(items_liste) and not defectueux
        return (
            f"<tr><td class='bold'><span style='display:inline-flex;align-items:center;gap:8px;'>"
            f"{icone_badge(icone_svg)}<span>{nom}</span></span></td>"
            f"<td class='center'>{case(conforme, '#16a34a')}</td>"
            f"<td class='center'>{case(defectueux, '#e11324')}</td>"
            f"<td class='center'>{case(so, '#9ca3af')}</td>"
            f"<td class='center'>{_badge_equipement(conforme, defectueux, so)}</td></tr>"
        )

    equipement_rows = (
        _ligne_equipement(t("systeme_cuisine"), ICONE_CUISINE, False, [])
        + _ligne_equipement(t("extincteur_label"), ICONE_EXTINCTEUR, True, items)
        + _ligne_equipement(t("eclairage_urgence_label"), ICONE_SORTIE, rapport_eclairage is not None, eclairages)
    )

    logo_content = organisation_logo_content(bat.client.organisation, 46)
    organisation_nom = bat.client.organisation.nom
    emetteur = cert.emis_par.get_full_name() or cert.emis_par.username if cert.emis_par else "—"

    entete_html = entete(
        logo_content, organisation_nom, f"{t('inspection_certification')} — {t('extincteurs_portatifs')}",
        t("certificat_no"), cert.numero, t("date_inspection"), date_insp, t("technicien_s"), tech_noms,
    )

    return f"""<!DOCTYPE html>
<html lang="{langue}">
<head>
<meta charset="UTF-8">
<title>{t("certificat_no")} {cert.numero}</title>
<style>{CSS_DOCUMENT}</style>
</head>
<body>
<div class="no-print" style="text-align:right;padding:8px 12px;background:#f8fafc;border-bottom:1px solid #e5e7eb;">
  <button onclick="window.print()" style="background:#0a0b0d;color:#fff;border:none;padding:8px 20px;border-radius:4px;font-weight:700;cursor:pointer;font-size:10pt;">{t("imprimer_pdf")}</button>
</div>
<div style="padding:20px 24px;">
{entete_html}
<div class="title-banner">
  <h2>{t("certificat_verification")}</h2>
  <div style="width:140px;height:1.5px;background:linear-gradient(90deg, transparent, #e11324, transparent);margin:6px auto;"></div>
  <p>{t("extincteurs_portatifs")}</p>
</div>
<div style="text-align:center;margin-bottom:14px;">
  <span style="display:inline-block;background:{conformite_bg};border:1.5px solid {conformite_color};color:{conformite_color};font-size:11pt;font-weight:900;letter-spacing:2px;padding:5px 22px;border-radius:100px;">{t("conforme_badge") if est_conforme else t("non_conforme_badge")}</span>
</div>
<div style="text-align:left;margin-bottom:6px;">
  <div class="card-title">{t("client")}</div>
  <div class="card-main" style="font-size:11pt;">{bat.client.nom}</div>
</div>
<div style="text-align:center;margin-bottom:6px;">
  <div class="card-title">{t("adresse_inspectee")}</div>
</div>
<div class="info-card" style="display:flex;align-items:center;justify-content:center;gap:14px;margin-bottom:18px;">
  <span style="display:inline-flex;align-items:center;justify-content:center;width:36px;height:36px;border-radius:50%;background:#f1f5f9;flex-shrink:0;">{icone(ICONE_PIN, 16, '#6b7280')}</span>
  <div class="card-main" style="font-size:20pt; font-weight:900;">{adresse}</div>
</div>
<div style="background:#0a0b0d;color:#fff;text-align:center;padding:7px 10px;border-radius:4px;margin-bottom:8px;">
  <span style="display:inline-flex;align-items:center;gap:6px;font-size:8pt;font-weight:800;letter-spacing:0.3px;">{icone(ICONE_BOUCLIER, 13, '#fff')}{t("conformite_bandeau")}</span>
</div>
<table class="equip-table">
  <thead><tr><th>{t("equipement")}</th><th class="center">{t("conforme_col")}</th><th class="center">{t("non_conforme_col")}</th><th class="center">{t("so")}</th><th class="center">{t("statut_col")}</th></tr></thead>
  <tbody>{equipement_rows}</tbody>
</table>
<p style="text-align:center;font-weight:700;font-size:8.5pt;color:#0a0b0d;margin-top:14px;line-height:1.4;">
  {t("inspection_entretien")}
</p>
<div class="sig-row">
  <div class="sig-block" style="display:flex;align-items:center;gap:10px;">
    <span class="sig-icon">{icone(ICONE_PERSONNE, 14, '#e11324')}</span>
    <div>
      <div class="sig-label">{t("superviseur_responsable")}</div>
      <div class="sig-name">{emetteur}</div>
      <div style="font-size:8pt;color:#555;">{organisation_nom}</div>
    </div>
  </div>
  <div class="sig-block" style="display:flex;align-items:center;gap:10px;">
    <span class="sig-icon">{icone(ICONE_CALENDRIER, 14, '#e11324')}</span>
    <div>
      <div class="sig-label">{t("date_emission")}</div>
      <div class="sig-name">{date_cert}</div>
      <div style="font-size:8pt;color:#555;">{t("certificat_no")} {cert.numero}</div>
    </div>
  </div>
</div>
{pied_de_page(organisation_nom, t("footer_certificat_extincteur"))}
</div>
</body>
</html>"""


def _html_rapport_extincteur_complet(rapport) -> str:
    """HTML du rapport technique complet de vérification des extincteurs
    portatifs — même chrome que le certificat, inclut l'éclairage d'urgence
    lié le cas échéant."""
    from .pdf_design import CSS_DOCUMENT, entete, pied_de_page

    bat = rapport.batiment
    adresse = f"{bat.numero_civique} {bat.rue}, {bat.ville}"
    langue = bat.client.organisation.langue
    t = lambda cle: _t(langue, cle)

    date_insp = _date_fr(rapport.date_inspection)
    techniciens = list(rapport.techniciens.all())
    tech_noms = ", ".join(t2.get_full_name() or t2.username for t2 in techniciens) or "—"

    legende_rows = "".join(
        f"<tr><td class='bold' style='width:50px;'>{code}</td><td>{desc}</td></tr>"
        for code, desc in LEGENDE_EXTINCTEURS
    )

    items = list(rapport.extincteurs.all())
    item_rows = ""
    for it in items:
        is_defect = it.etat == ExtincteurItem.Etat.DEFECTUEUX
        is_ni = not is_defect and it.etat == "NI"
        bg = ' style="background:#fef2f2;"' if is_defect else ' style="background:#fef3c7;"' if is_ni else ""
        etat_style = ' style="color:#cc0000;"' if is_defect else ' style="color:#b45309;"' if is_ni else ""
        item_rows += (
            f"<tr{bg}>"
            f"<td class='center'>{it.ordre}</td>"
            f"<td>{it.etage or '—'}</td>"
            f"<td>{it.emplacement or '—'}</td>"
            f"<td class='center'>{_td(langue, 'type_extincteur', it.type_extincteur) or '—'}</td>"
            f"<td class='center'>{_td(langue, 'format', it.format) or '—'}</td>"
            f"<td>{_td(langue, 'marque', it.marque) or '—'}</td>"
            f"<td>{it.numero_serie or '—'}</td>"
            f"<td class='center'>{it.date_fabrication or '—'}</td>"
            f"<td class='center'>{it.prochaine_maintenance or '—'}</td>"
            f"<td class='center'>{it.prochain_test_hydrostatique or '—'}</td>"
            f"<td class='center bold'{etat_style}>{it.etat or '—'}</td>"
            f"<td>{it.remarque or ''}</td>"
            f"</tr>"
        )
    if not item_rows:
        item_rows = f"<tr><td colspan='12' class='muted center'>{t('aucun_extincteur')}</td></tr>"

    boyaux = list(rapport.boyaux.all())
    boyau_rows = ""
    for b in boyaux:
        is_defect = b.etat == BoyauItem.Etat.DEFECTUEUX
        is_ni = not is_defect and b.etat == "NI"
        bg = ' style="background:#fef2f2;"' if is_defect else ' style="background:#fef3c7;"' if is_ni else ""
        etat_style = ' style="color:#cc0000;"' if is_defect else ' style="color:#b45309;"' if is_ni else ""
        boyau_rows += (
            f"<tr{bg}>"
            f"<td class='center'>{b.ordre}</td>"
            f"<td>{b.etage or '—'}</td>"
            f"<td class='center bold'{etat_style}>{b.etat or '—'}</td>"
            f"<td>{b.emplacement or '—'}</td>"
            f"<td class='center'>{_td(langue, 'longueur', b.longueur) or '—'}</td>"
            f"<td class='center'>{b.date_fabrication or '—'}</td>"
            f"<td class='center'>{b.prochain_test_hydrostatique or '—'}</td>"
            f"<td>{b.remarque or ''}</td>"
            f"</tr>"
        )
    if not boyau_rows:
        boyau_rows = f"<tr><td colspan='8' class='muted center'>{t('aucun_boyau')}</td></tr>"

    # ── Éclairage d'urgence lié (même visite, voir rapport_eclairage_lie) ──
    rapport_eclairage = getattr(rapport, "rapport_eclairage_lie", None)
    eclairages = list(rapport_eclairage.eclairages_urgence.all()) if rapport_eclairage else []
    eclairage_section = ""
    if eclairages:
        eclairage_rows = ""
        for it in eclairages:
            is_defect = it.etat == "D"
            is_ni = not is_defect and it.etat == "NI"
            bg = ' style="background:#fef2f2;"' if is_defect else ' style="background:#fef3c7;"' if is_ni else ""
            etat_style = ' style="color:#cc0000;"' if is_defect else ' style="color:#b45309;"' if is_ni else ""
            eclairage_rows += (
                f"<tr{bg}>"
                f"<td class='center'>{it.ordre}</td>"
                f"<td>{it.etage or '—'}</td>"
                f"<td>{it.emplacement or '—'}</td>"
                f"<td>{it.modele or '—'}</td>"
                f"<td class='center'>{it.voltage or '—'}</td>"
                f"<td class='center bold'{etat_style}>{it.etat or '—'}</td>"
                f"<td>{it.remarque or ''}</td>"
                f"</tr>"
            )
        eclairage_section = f"""<div class="sec-title">{t("detail_unites_eclairage")}</div>
<table>
  <thead><tr>
    <th>{t("col_no")}</th><th>{t("col_etage")}</th><th>{t("col_emplacement")}</th><th>{t("col_modele")}</th>
    <th>{t("col_voltage")}</th><th title="{t('etat_titre_abbr')}">{t("col_etat")}</th><th>{t("col_remarque")}</th>
  </tr></thead>
  <tbody>{eclairage_rows}</tbody>
</table>"""

    logo_content = organisation_logo_content(bat.client.organisation, 46)
    organisation_nom = bat.client.organisation.nom

    entete_html = entete(
        logo_content, organisation_nom, f"{t('rapport_verification')} — {t('extincteurs_portatifs')}",
        t("certificat_no"), (rapport.certificat.numero if hasattr(rapport, "certificat") else "—"),
        t("date_inspection"), date_insp, t("technicien_s"), tech_noms,
    )

    return f"""<!DOCTYPE html>
<html lang="{langue}">
<head>
<meta charset="UTF-8">
<title>{t("rapport_verification")} — {t("extincteurs_portatifs")} — {adresse}</title>
<style>{CSS_DOCUMENT}</style>
</head>
<body>
<div class="no-print" style="text-align:right;padding:8px 12px;background:#f8fafc;border-bottom:1px solid #e5e7eb;">
  <button onclick="window.print()" style="background:#0a0b0d;color:#fff;border:none;padding:8px 20px;border-radius:4px;font-weight:700;cursor:pointer;font-size:10pt;">{t("imprimer_pdf")}</button>
</div>
<div style="padding:16px 20px;">
{entete_html}
<div class="title-banner">
  <h2>{t("rapport_verification")} — {t("extincteurs_portatifs")}</h2>
  <div style="width:140px;height:1.5px;background:linear-gradient(90deg, transparent, #e11324, transparent);margin:6px auto;"></div>
</div>
<div style="text-align:left;margin-bottom:6px;">
  <div class="card-title">{t("client")}</div>
  <div class="card-main" style="font-size:10.5pt;">{bat.client.nom}</div>
</div>
<div class="info-card" style="text-align:center;margin-bottom:18px;">
  <div class="card-title">{t("adresse")}</div>
  <div class="card-main" style="font-size:14pt;">{adresse}</div>
</div>
<div class="legende-box">
<table><tbody>{legende_rows}</tbody></table>
</div>
<div class="sec-title">{t("detail_extincteurs")}</div>
<table>
  <thead><tr>
    <th>{t("col_no")}</th><th>{t("col_etage")}</th><th>{t("col_emplacement")}</th><th>{t("col_type")}</th><th>{t("col_format")}</th><th>{t("col_marque")}</th><th>{t("col_numero_serie")}</th>
    <th>{t("col_date_fabrication")}</th><th>{t("col_prochaine_maintenance")}</th><th>{t("col_prochain_test_hydro")}</th>
    <th title="{t('etat_titre_abbr')}">{t("col_etat")}</th><th>{t("col_remarque")}</th>
  </tr></thead>
  <tbody>{item_rows}</tbody>
</table>
<div class="sec-title">{t("detail_boyaux")}</div>
<table>
  <thead><tr>
    <th>{t("col_no")}</th><th>{t("col_etage")}</th><th title="{t('etat_titre_abbr')}">{t("col_etat")}</th><th>{t("col_emplacement")}</th>
    <th>{t("col_longueur")}</th><th>{t("col_annee_fabrication")}</th><th>{t("col_prochain_test_hydro")}</th><th>{t("col_remarque")}</th>
  </tr></thead>
  <tbody>{boyau_rows}</tbody>
</table>
{eclairage_section}
{pied_de_page(organisation_nom, t("footer_rapport_extincteur"))}
</div>
</body>
</html>"""


# ── Rapport extincteurs portatifs ────────────────────────────────────────
class EstModuleRapportExtincteurActif(permissions.BasePermission):
    message = "Le module « Rapport extincteur » n'est pas activé pour votre organisation."

    def has_permission(self, request, view):
        organisation = getattr(request.user, "organisation", None)
        return bool(organisation and organisation.a_le_module("rapport_extincteur"))


class RapportExtincteurViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated, EstModuleRapportExtincteurActif]

    def get_serializer_class(self):
        if self.action == "list":
            return RapportExtincteurListSerializer
        if self.action in ["create", "update", "partial_update"]:
            # Le superviseur peut réassigner les techniciens (ex. absence) —
            # RapportExtincteurDetailSerializer déclare `techniciens` en lecture
            # seule (affichage imbriqué), il faut le serializer d'écriture ici.
            return RapportExtincteurCreateSerializer
        return RapportExtincteurDetailSerializer

    def get_queryset(self):
        user = self.request.user
        qs = RapportExtincteur.objects.select_related(
            "batiment", "batiment__client", "cree_par", "citoyen"
        ).prefetch_related("techniciens").filter(batiment__client__organisation=user.organisation)

        if user.est_citoyen():
            qs = qs.filter(citoyen=user)
        elif user.est_technicien():
            qs = qs.filter(techniciens=user)
        # le superviseur voit tout

        client_id = self.request.query_params.get("client")
        statut = self.request.query_params.get("statut")
        if client_id:
            qs = qs.filter(batiment__client_id=client_id)
        if statut:
            qs = qs.filter(statut=statut)

        return qs.distinct()

    def get_permissions(self):
        if self.action in ["create", "destroy", "update", "partial_update", "reassigner", "rouvrir"]:
            # Contrairement au rapport principal, le technicien n'a jamais besoin
            # d'écrire directement sur l'objet RapportExtincteur (il modifie les
            # lignes via l'action `extincteurs` / ExtincteurItemViewSet) — donc
            # batiment/techniciens/citoyen restent réservés au superviseur.
            return [permissions.IsAuthenticated(), EstSuperviseur()]
        return super().get_permissions()

    @action(detail=True, methods=["patch"])
    def reassigner(self, request, pk=None):
        """Superviseur seulement — change le bâtiment et/ou les techniciens assignés après création."""
        rapport = self.get_object()
        changements = []

        if "batiment" in request.data:
            try:
                nouveau_batiment = Batiment.objects.get(pk=request.data["batiment"])
            except (Batiment.DoesNotExist, TypeError, ValueError):
                return Response({"error": "Bâtiment introuvable."}, status=status.HTTP_400_BAD_REQUEST)
            if nouveau_batiment.id != rapport.batiment_id:
                rapport.batiment = nouveau_batiment
                changements.append(f"Bâtiment changé pour {nouveau_batiment.adresse_complete}")

        if "techniciens" in request.data:
            rapport.techniciens.set(request.data.get("techniciens") or [])
            changements.append("Techniciens réassignés")

        if "citoyen" in request.data:
            citoyen_id = request.data.get("citoyen")
            if citoyen_id:
                try:
                    nouveau_citoyen = Utilisateur.objects.get(pk=citoyen_id, role=Utilisateur.Role.CITOYEN)
                except (Utilisateur.DoesNotExist, TypeError, ValueError):
                    return Response({"error": "Citoyen introuvable."}, status=status.HTTP_400_BAD_REQUEST)
                rapport.citoyen = nouveau_citoyen
                changements.append(f"Citoyen réassigné à {nouveau_citoyen.username}")
            else:
                rapport.citoyen = None
                changements.append("Citoyen retiré du rapport")

        if changements:
            rapport.save()
            for c in changements:
                rapport.historiser(request.user, c)

        return Response(RapportExtincteurDetailSerializer(rapport).data)

    @action(detail=True, methods=["post"])
    def rouvrir(self, request, pk=None):
        rapport = self.get_object()
        if not request.user.est_superviseur():
            return Response(
                {"error": "Seul le superviseur peut rouvrir un rapport."},
                status=status.HTTP_403_FORBIDDEN,
            )
        if rapport.statut != RapportExtincteur.Statut.FERME:
            return Response({"error": "Ce rapport est déjà ouvert."}, status=status.HTTP_400_BAD_REQUEST)

        rapport.rouvrir(request.user)
        return Response(RapportExtincteurDetailSerializer(rapport).data)

    def perform_create(self, serializer):
        rapport = serializer.save(cree_par=self.request.user)
        rapport.historiser(self.request.user, "Rapport créé")

        # Une inspection couvre extincteurs + éclairage d'urgence en même
        # temps — le rapport éclairage correspondant est donc créé et lié
        # automatiquement, pour n'avoir qu'un seul certificat à la fermeture.
        _creer_rapport_eclairage_lie(rapport, self.request.user)

    def perform_update(self, serializer):
        instance = self.get_object()
        if instance.statut == RapportExtincteur.Statut.FERME and not self.request.user.est_superviseur():
            from rest_framework.exceptions import ValidationError
            raise ValidationError("Ce rapport est fermé et ne peut plus être modifié.")
        rapport = serializer.save()
        rapport.historiser(self.request.user, "Rapport modifié")

    @action(detail=True, methods=["post"])
    def fermer(self, request, pk=None):
        rapport = self.get_object()
        if not request.user.est_superviseur():
            return Response(
                {"error": "Seul le superviseur peut fermer un rapport."},
                status=status.HTTP_403_FORBIDDEN,
            )
        if rapport.statut == RapportExtincteur.Statut.FERME:
            return Response({"error": "Ce rapport est déjà fermé."}, status=status.HTTP_400_BAD_REQUEST)

        rapport.fermer(request.user)
        return Response(RapportExtincteurDetailSerializer(rapport).data)

    @action(detail=True, methods=["post"], url_path="envoyer-certificat")
    def envoyer_certificat(self, request, pk=None):
        rapport = self.get_object()
        if not request.user.est_superviseur():
            return Response(
                {"error": "Seul le superviseur peut envoyer le certificat."},
                status=status.HTTP_403_FORBIDDEN,
            )
        if rapport.statut != RapportExtincteur.Statut.FERME:
            return Response(
                {"error": "Le rapport doit être fermé avant d'envoyer le certificat."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if not hasattr(rapport, "certificat"):
            return Response({"error": "Aucun certificat trouvé pour ce rapport."}, status=status.HTTP_404_NOT_FOUND)

        client = rapport.batiment.client
        if client.mode_livraison == Client.ModeLivraison.DIRECT:
            from .emailing import envoyer_certificats_directs_batiment

            ok, message = envoyer_certificats_directs_batiment(rapport.batiment, request.user)
            if not ok:
                return Response({"error": message}, status=status.HTTP_400_BAD_REQUEST)
            return Response({"message": message})

        rapport.certificat.certificat_envoye = True
        rapport.certificat.save()
        rapport.historiser(request.user, f"Certificat envoyé au citoyen {rapport.citoyen.username if rapport.citoyen else '—'}")

        if rapport.citoyen and rapport.citoyen.email:
            from .emailing import envoyer_email_certificat_extincteur_disponible

            envoyer_email_certificat_extincteur_disponible(rapport)

        return Response({"message": "Certificat envoyé au citoyen."})

    @action(detail=True, methods=["get", "post"])
    def extincteurs(self, request, pk=None):
        rapport = self.get_object()

        if request.method == "GET":
            return Response(ExtincteurItemSerializer(rapport.extincteurs.all(), many=True).data)

        if rapport.statut == RapportExtincteur.Statut.FERME and not request.user.est_superviseur():
            return Response(
                {"error": "Ce rapport est fermé, impossible d'ajouter un extincteur."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        serializer = ExtincteurItemSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        ordre = serializer.validated_data.get("ordre") or (rapport.extincteurs.count() + 1)
        serializer.save(rapport=rapport, ordre=ordre)
        rapport.historiser(request.user, "Extincteur ajouté")
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=["get", "post"])
    def boyaux(self, request, pk=None):
        rapport = self.get_object()

        if request.method == "GET":
            return Response(BoyauItemSerializer(rapport.boyaux.all(), many=True).data)

        if rapport.statut == RapportExtincteur.Statut.FERME and not request.user.est_superviseur():
            return Response(
                {"error": "Ce rapport est fermé, impossible d'ajouter un boyau."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        serializer = BoyauItemSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        ordre = serializer.validated_data.get("ordre") or (rapport.boyaux.count() + 1)
        serializer.save(rapport=rapport, ordre=ordre)
        rapport.historiser(request.user, "Boyau ajouté")
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=["get"])
    def historique(self, request, pk=None):
        rapport = self.get_object()
        return Response(HistoriqueRapportExtincteurSerializer(rapport.historique.all(), many=True).data)

    @action(detail=True, methods=["get"], url_path="certificat-pdf")
    def certificat_pdf(self, request, pk=None):
        rapport = self.get_object()
        if rapport.statut != RapportExtincteur.Statut.FERME:
            return Response({"error": "Le rapport doit être fermé."}, status=status.HTTP_400_BAD_REQUEST)
        if not hasattr(rapport, "certificat"):
            return Response({"error": "Aucun certificat pour ce rapport."}, status=status.HTTP_404_NOT_FOUND)

        html = _html_certificat_extincteur(rapport)
        return HttpResponse(html, content_type="text/html; charset=utf-8")

    @action(detail=True, methods=["get"], url_path="telecharger")
    def telecharger(self, request, pk=None):
        rapport = self.get_object()
        html = _html_rapport_extincteur_complet(rapport)
        return HttpResponse(html, content_type="text/html; charset=utf-8")

    @action(detail=True, methods=["get"])
    def excel(self, request, pk=None):
        rapport = self.get_object()
        bat = rapport.batiment
        adresse = f"{bat.numero_civique} {bat.rue}, {bat.ville}"
        langue = bat.client.organisation.langue
        t = lambda cle: _t(langue, cle)
        techniciens = ", ".join(
            t2.get_full_name() or t2.username for t2 in rapport.techniciens.all()
        ) or "—"

        wb = excel_workbook()
        ws = wb.active
        ws.title = t("extincteur_label")
        excel_entete_rapport(
            ws, organisation_nom=bat.client.organisation.nom, adresse=adresse,
            date_insp=_date_fr(rapport.date_inspection),
            statut=t('statut_ferme') if rapport.statut == 'ferme' else t('statut_ouvert'), techniciens=techniciens,
        )

        colonnes = [
            t("col_no"), t("col_etage"), t("col_emplacement"), t("col_type"), t("col_format"), t("col_marque"), t("col_numero_serie"),
            t("col_date_fabrication"), t("col_prochaine_maintenance"), t("col_prochain_test_hydro"),
            t("col_etat"), t("col_remarque"),
        ]
        ligne = excel_ligne_entetes(ws, colonnes, ligne=6)
        for it in rapport.extincteurs.all():
            ligne += 1
            valeurs = [
                it.ordre, it.etage, it.emplacement,
                _td(langue, 'type_extincteur', it.type_extincteur),
                _td(langue, 'format', it.format),
                _td(langue, 'marque', it.marque), it.numero_serie,
                it.date_fabrication, it.prochaine_maintenance, it.prochain_test_hydrostatique,
                it.etat, it.remarque,
            ]
            for col, valeur in enumerate(valeurs, start=1):
                ws.cell(row=ligne, column=col, value=valeur)
        excel_ajuster_largeurs(ws, colonnes)
        ws.freeze_panes = "A7"

        ws2 = wb.create_sheet(t("boyaux_incendie_sheet"))
        excel_entete_rapport(
            ws2, organisation_nom=bat.client.organisation.nom, adresse=adresse,
            date_insp=_date_fr(rapport.date_inspection),
            statut=t('statut_ferme') if rapport.statut == 'ferme' else t('statut_ouvert'), techniciens=techniciens,
        )
        colonnes_boyaux = [
            t("col_no"), t("col_etage"), t("col_etat"), t("col_emplacement"), t("col_longueur"),
            t("col_annee_fabrication"), t("col_prochain_test_hydro"), t("col_remarque"),
        ]
        ligne2 = excel_ligne_entetes(ws2, colonnes_boyaux, ligne=6)
        for b in rapport.boyaux.all():
            ligne2 += 1
            valeurs2 = [
                b.ordre, b.etage, b.etat, b.emplacement,
                _td(langue, 'longueur', b.longueur),
                b.date_fabrication, b.prochain_test_hydrostatique, b.remarque,
            ]
            for col, valeur in enumerate(valeurs2, start=1):
                ws2.cell(row=ligne2, column=col, value=valeur)
        excel_ajuster_largeurs(ws2, colonnes_boyaux)
        ws2.freeze_panes = "A7"

        return excel_reponse(wb, excel_nom_fichier(t("extincteur_label"), adresse))


class ExtincteurItemViewSet(viewsets.ModelViewSet):
    """Accès direct à une ligne d'extincteur — pour la corriger ou la supprimer."""

    serializer_class = ExtincteurItemSerializer
    permission_classes = [permissions.IsAuthenticated, EstSuperviseurOuTechnicien, EstModuleRapportExtincteurActif]

    def get_queryset(self):
        user = self.request.user
        qs = ExtincteurItem.objects.select_related("rapport").filter(rapport__batiment__client__organisation=user.organisation)
        if user.est_technicien():
            qs = qs.filter(rapport__techniciens=user)
        return qs.distinct()

    def perform_update(self, serializer):
        item = self.get_object()
        if item.rapport.statut == RapportExtincteur.Statut.FERME and not self.request.user.est_superviseur():
            from rest_framework.exceptions import ValidationError
            raise ValidationError("Le rapport associé est fermé.")
        serializer.save()


class BoyauItemViewSet(viewsets.ModelViewSet):
    """Accès direct à une ligne de boyau — pour la corriger ou la supprimer."""

    serializer_class = BoyauItemSerializer
    permission_classes = [permissions.IsAuthenticated, EstSuperviseurOuTechnicien, EstModuleRapportExtincteurActif]

    def get_queryset(self):
        user = self.request.user
        qs = BoyauItem.objects.select_related("rapport").filter(rapport__batiment__client__organisation=user.organisation)
        if user.est_technicien():
            qs = qs.filter(rapport__techniciens=user)
        return qs.distinct()

    def perform_update(self, serializer):
        item = self.get_object()
        if item.rapport.statut == RapportExtincteur.Statut.FERME and not self.request.user.est_superviseur():
            from rest_framework.exceptions import ValidationError
            raise ValidationError("Le rapport associé est fermé.")
        serializer.save()


class EstModuleRapportEclairageUrgenceActif(permissions.BasePermission):
    message = "Le module « Rapport éclairage d'urgence » n'est pas activé pour votre organisation."

    def has_permission(self, request, view):
        organisation = getattr(request.user, "organisation", None)
        return bool(organisation and organisation.a_le_module("rapport_eclairage_urgence"))


class RapportEclairageUrgenceViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated, EstModuleRapportEclairageUrgenceActif]

    def get_serializer_class(self):
        if self.action == "list":
            return RapportEclairageUrgenceListSerializer
        if self.action in ["create", "update", "partial_update"]:
            return RapportEclairageUrgenceCreateSerializer
        return RapportEclairageUrgenceDetailSerializer

    def get_queryset(self):
        user = self.request.user
        qs = RapportEclairageUrgence.objects.select_related(
            "batiment", "batiment__client", "cree_par"
        ).prefetch_related("techniciens").filter(batiment__client__organisation=user.organisation)

        if user.est_technicien():
            qs = qs.filter(techniciens=user)
        # le superviseur voit tout

        client_id = self.request.query_params.get("client")
        statut = self.request.query_params.get("statut")
        if client_id:
            qs = qs.filter(batiment__client_id=client_id)
        if statut:
            qs = qs.filter(statut=statut)

        return qs.distinct()

    def get_permissions(self):
        if self.action in ["create", "destroy", "update", "partial_update", "reassigner", "rouvrir"]:
            return [permissions.IsAuthenticated(), EstSuperviseur()]
        return super().get_permissions()

    @action(detail=True, methods=["patch"])
    def reassigner(self, request, pk=None):
        """Superviseur seulement — change le bâtiment et/ou les techniciens assignés après création."""
        rapport = self.get_object()
        changements = []

        if "batiment" in request.data:
            try:
                nouveau_batiment = Batiment.objects.get(pk=request.data["batiment"])
            except (Batiment.DoesNotExist, TypeError, ValueError):
                return Response({"error": "Bâtiment introuvable."}, status=status.HTTP_400_BAD_REQUEST)
            if nouveau_batiment.id != rapport.batiment_id:
                rapport.batiment = nouveau_batiment
                changements.append(f"Bâtiment changé pour {nouveau_batiment.adresse_complete}")

        if "techniciens" in request.data:
            rapport.techniciens.set(request.data.get("techniciens") or [])
            changements.append("Techniciens réassignés")

        if changements:
            rapport.save()
            for c in changements:
                rapport.historiser(request.user, c)

        return Response(RapportEclairageUrgenceDetailSerializer(rapport).data)

    @action(detail=True, methods=["post"])
    def rouvrir(self, request, pk=None):
        rapport = self.get_object()
        if not request.user.est_superviseur():
            return Response(
                {"error": "Seul le superviseur peut rouvrir un rapport."},
                status=status.HTTP_403_FORBIDDEN,
            )
        if rapport.statut != RapportEclairageUrgence.Statut.FERME:
            return Response({"error": "Ce rapport est déjà ouvert."}, status=status.HTTP_400_BAD_REQUEST)

        rapport.rouvrir(request.user)
        return Response(RapportEclairageUrgenceDetailSerializer(rapport).data)

    def perform_create(self, serializer):
        rapport = serializer.save(cree_par=self.request.user)
        rapport.historiser(self.request.user, "Rapport créé")

    def perform_update(self, serializer):
        instance = self.get_object()
        if instance.statut == RapportEclairageUrgence.Statut.FERME and not self.request.user.est_superviseur():
            from rest_framework.exceptions import ValidationError
            raise ValidationError("Ce rapport est fermé et ne peut plus être modifié.")
        rapport = serializer.save()
        rapport.historiser(self.request.user, "Rapport modifié")

    @action(detail=True, methods=["post"])
    def fermer(self, request, pk=None):
        rapport = self.get_object()
        if not request.user.est_superviseur():
            return Response(
                {"error": "Seul le superviseur peut fermer un rapport."},
                status=status.HTTP_403_FORBIDDEN,
            )
        if rapport.statut == RapportEclairageUrgence.Statut.FERME:
            return Response({"error": "Ce rapport est déjà fermé."}, status=status.HTTP_400_BAD_REQUEST)

        rapport.fermer(request.user)
        return Response(RapportEclairageUrgenceDetailSerializer(rapport).data)

    @action(detail=True, methods=["get", "post"], url_path="eclairages-urgence")
    def eclairages_urgence(self, request, pk=None):
        rapport = self.get_object()

        if request.method == "GET":
            return Response(EclairageUrgenceItemSerializer(rapport.eclairages_urgence.all(), many=True).data)

        if rapport.statut == RapportEclairageUrgence.Statut.FERME and not request.user.est_superviseur():
            return Response(
                {"error": "Ce rapport est fermé, impossible d'ajouter une unité."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        serializer = EclairageUrgenceItemSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        ordre = serializer.validated_data.get("ordre") or (rapport.eclairages_urgence.count() + 1)
        serializer.save(rapport=rapport, ordre=ordre)
        rapport.historiser(request.user, "Unité d'éclairage ajoutée")
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=["get"])
    def historique(self, request, pk=None):
        rapport = self.get_object()
        return Response(HistoriqueRapportEclairageUrgenceSerializer(rapport.historique.all(), many=True).data)

    @action(detail=True, methods=["get"], url_path="telecharger")
    def telecharger(self, request, pk=None):
        rapport = self.get_object()
        bat = rapport.batiment
        adresse = f"{bat.numero_civique} {bat.rue}, {bat.ville}"
        from django.utils import timezone

        langue = bat.client.organisation.langue
        t = lambda cle: _t(langue, cle)

        date_insp = _date_fr(timezone.localdate())
        techniciens = list(rapport.techniciens.all())
        tech_noms = ", ".join(t2.get_full_name() or t2.username for t2 in techniciens) or "—"

        items = list(rapport.eclairages_urgence.all())
        item_rows = ""
        for it in items:
            is_defect = it.etat == EclairageUrgenceItem.Etat.DEFECTUEUX
            is_ni = not is_defect and it.etat == "NI"
            bg = ' style="background:#fef2f2;"' if is_defect else ' style="background:#fef3c7;"' if is_ni else ""
            etat_style = ' style="color:#cc0000;"' if is_defect else ' style="color:#b45309;"' if is_ni else ""
            item_rows += (
                f"<tr{bg}>"
                f"<td class='center'>{it.ordre}</td>"
                f"<td>{it.emplacement or '—'}</td>"
                f"<td>{it.etage or '—'}</td>"
                f"<td>{it.modele or '—'}</td>"
                f"<td>{it.voltage or '—'}</td>"
                f"<td class='center bold'{etat_style}>{it.etat or '—'}</td>"
                f"<td>{it.remarque or ''}</td>"
                f"</tr>"
            )
        if not item_rows:
            item_rows = f"<tr><td colspan='7' class='muted center'>{t('aucune_unite')}</td></tr>"

        logo_content = organisation_logo_content(bat.client.organisation, 46)
        organisation_nom = bat.client.organisation.nom

        html = f"""<!DOCTYPE html>
<html lang="{langue}">
<head>
<meta charset="UTF-8">
<title>{t("footer_rapport_eclairage")} — {adresse}</title>
<style>
  @page {{ margin: 14mm 12mm; }}
  *{{ box-sizing:border-box; margin:0; padding:0; }}
  body{{ font-family:Arial,Helvetica,sans-serif; font-size:9pt; color:#000; background:#fff; }}
  .header{{ display:flex; align-items:center; justify-content:space-between; border-bottom:3px solid #0a0b0d; padding-bottom:10px; margin-bottom:14px; }}
  .brand{{ display:flex; align-items:center; gap:12px; }}
  .logo-box{{ height:46px; max-width:150px; display:flex; align-items:center; flex-shrink:0; }}
  .logo-box img{{ max-height:100%; max-width:100%; }}
  .brand-text h1{{ font-size:12pt; font-weight:900; color:#0a0b0d; text-transform:uppercase; }}
  .brand-text p{{ font-size:7.5pt; color:#444; margin-top:1px; }}
  .info-card{{ border:1px solid #ccc; border-radius:4px; padding:8px 12px; }}
  .card-title{{ font-size:7pt; font-weight:700; text-transform:uppercase; letter-spacing:1.5px; color:#555; margin-bottom:4px; }}
  .card-main{{ font-size:10pt; font-weight:700; color:#000; }}
  .title-banner{{ background:#0a0b0d; color:#fff; text-align:center; padding:8px 0; border-radius:4px; margin-bottom:12px; }}
  .title-banner h2{{ font-size:11pt; font-weight:700; letter-spacing:1.5px; text-transform:uppercase; }}
  .sec-title{{ font-size:8.5pt; font-weight:700; text-transform:uppercase; color:#000; border-bottom:2px solid #000; padding-bottom:3px; margin-bottom:6px; margin-top:14px; }}
  table{{ width:100%; border-collapse:collapse; font-size:8pt; }}
  th{{ background:#e8e8e8; color:#000; font-weight:700; padding:4px 6px; text-align:left; font-size:7.5pt; border:1px solid #ccc; }}
  td{{ padding:4px 6px; border:1px solid #ddd; color:#000; }}
  .center{{ text-align:center; }} .bold{{ font-weight:700; }} .muted{{ color:#777; font-style:italic; }}
  .footer{{ margin-top:20px; padding-top:8px; border-top:1px solid #ccc; display:flex; justify-content:space-between; font-size:7pt; color:#555; }}
  @media print{{ body{{ -webkit-print-color-adjust:exact; print-color-adjust:exact; }} .no-print{{ display:none!important; }} }}
</style>
</head>
<body>
<div class="no-print" style="text-align:right;padding:8px 12px;background:#f8fafc;border-bottom:1px solid #e5e7eb;">
  <button onclick="window.print()" style="background:#0a0b0d;color:#fff;border:none;padding:8px 20px;border-radius:4px;font-weight:700;cursor:pointer;font-size:10pt;">{t("imprimer_pdf")}</button>
</div>
<div style="padding:16px 20px;">
<div class="header">
  <div class="brand">
    <div class="logo-box">{logo_content}</div>
    <div class="brand-text">
      <h1>{organisation_nom}</h1>
      <p>{t("footer_rapport_eclairage")}</p>
    </div>
  </div>
  <div style="text-align:right;">
    <div style="font-size:8pt;font-weight:700;text-transform:uppercase;color:{'#0d6b4f' if rapport.statut == 'ferme' else '#ff6b1a'};">{t('statut_ferme') if rapport.statut == 'ferme' else t('statut_ouvert')}</div>
    <div style="font-size:7.5pt;color:#555;margin-top:4px;">{t("date_inspection")} : <strong style="color:#000;">{date_insp}</strong></div>
    <div style="font-size:7.5pt;color:#555;margin-top:1px;">{t("technicien_s")} : <strong style="color:#000;">{tech_noms}</strong></div>
  </div>
</div>
<div class="title-banner"><h2>{t("footer_rapport_eclairage")}</h2></div>
<div class="info-card" style="text-align:center;margin-bottom:18px;">
  <div class="card-title">{t("adresse")}</div>
  <div class="card-main" style="font-size:14pt;">{adresse}</div>
</div>
<div class="sec-title">{t("detail_unites_eclairage")}</div>
<table>
  <thead><tr>
    <th>{t("col_no")}</th><th>{t("col_emplacement")}</th><th>{t("col_etage")}</th><th>{t("col_modele")}</th><th>{t("col_voltage")}</th>
    <th title="{t('etat_titre_abbr')}">{t("col_etat")}</th><th>{t("col_remarque")}</th>
  </tr></thead>
  <tbody>{item_rows}</tbody>
</table>
<div class="footer">
  <div><strong>{organisation_nom}</strong></div>
  <div>{adresse}</div>
</div>
</div>
</body>
</html>"""
        return HttpResponse(html, content_type="text/html; charset=utf-8")

    @action(detail=True, methods=["get"])
    def excel(self, request, pk=None):
        rapport = self.get_object()
        bat = rapport.batiment
        adresse = f"{bat.numero_civique} {bat.rue}, {bat.ville}"
        langue = bat.client.organisation.langue
        t = lambda cle: _t(langue, cle)
        techniciens = ", ".join(
            t2.get_full_name() or t2.username for t2 in rapport.techniciens.all()
        ) or "—"

        wb = excel_workbook()
        ws = wb.active
        ws.title = t("eclairage_urgence_label")
        excel_entete_rapport(
            ws, organisation_nom=bat.client.organisation.nom, adresse=adresse,
            date_insp=_date_fr(rapport.date_inspection),
            statut=t('statut_ferme') if rapport.statut == 'ferme' else t('statut_ouvert'), techniciens=techniciens,
        )

        colonnes = [t("col_no"), t("col_emplacement"), t("col_etage"), t("col_modele"), t("col_voltage"), t("col_etat"), t("col_remarque")]
        ligne = excel_ligne_entetes(ws, colonnes, ligne=6)
        for it in rapport.eclairages_urgence.all():
            ligne += 1
            valeurs = [it.ordre, it.emplacement, it.etage, it.modele, it.voltage, it.etat, it.remarque]
            for col, valeur in enumerate(valeurs, start=1):
                ws.cell(row=ligne, column=col, value=valeur)
        excel_ajuster_largeurs(ws, colonnes)
        ws.freeze_panes = "A7"

        return excel_reponse(wb, excel_nom_fichier(t("eclairage_urgence_label"), adresse))


class EclairageUrgenceItemViewSet(viewsets.ModelViewSet):
    """Accès direct à une ligne d'appareil d'éclairage d'urgence — pour la corriger ou la supprimer."""

    serializer_class = EclairageUrgenceItemSerializer
    permission_classes = [permissions.IsAuthenticated, EstSuperviseurOuTechnicien, EstModuleRapportEclairageUrgenceActif]

    def get_queryset(self):
        user = self.request.user
        qs = EclairageUrgenceItem.objects.select_related("rapport").filter(rapport__batiment__client__organisation=user.organisation)
        if user.est_technicien():
            qs = qs.filter(rapport__techniciens=user)
        return qs.distinct()

    def perform_update(self, serializer):
        item = self.get_object()
        if item.rapport.statut == RapportEclairageUrgence.Statut.FERME and not self.request.user.est_superviseur():
            from rest_framework.exceptions import ValidationError
            raise ValidationError("Le rapport associé est fermé.")
        serializer.save()


# ── Certificats (vue unifiée, tous modules) ─────────────────────────────────

def _certificats_incendie(organisation):
    certs = Certificat.objects.select_related(
        "rapport", "rapport__batiment", "rapport__batiment__client", "emis_par"
    ).filter(rapport__batiment__client__organisation=organisation)
    resultats = []
    for c in certs:
        r = c.rapport
        bat = r.batiment
        resultats.append({
            "cle": f"incendie-{c.id}",
            "type": "incendie",
            "type_display": "Rapport incendie",
            "numero": c.numero,
            "date_emission": c.date_emission,
            "certificat_envoye": c.certificat_envoye,
            "conforme": c.conforme,
            "adresse": bat.adresse_complete,
            "client_nom": bat.client.nom,
            "client_id": bat.client_id,
            "rapport_id": r.id,
            "statut_rapport": r.statut,
            "url_rapport": f"/superviseur/rapports/{r.id}",
            "url_certificat_pdf": f"/api/rapports/{r.id}/certificat-pdf/",
        })
    return resultats


def _certificats_extincteur(organisation):
    certs = CertificatExtincteur.objects.select_related(
        "rapport", "rapport__batiment", "rapport__batiment__client", "emis_par"
    ).filter(rapport__batiment__client__organisation=organisation)
    resultats = []
    for c in certs:
        r = c.rapport
        bat = r.batiment
        resultats.append({
            "cle": f"extincteur-{c.id}",
            "type": "extincteur",
            "type_display": "Extincteur & éclairage",
            "numero": c.numero,
            "date_emission": c.date_emission,
            "certificat_envoye": c.certificat_envoye,
            "conforme": _est_conforme_extincteur(r),
            "adresse": bat.adresse_complete,
            "client_nom": bat.client.nom,
            "client_id": bat.client_id,
            "rapport_id": r.id,
            "statut_rapport": r.statut,
            "url_rapport": f"/superviseur/rapports-extincteurs/{r.id}",
            "url_certificat_pdf": f"/api/rapports-extincteurs/{r.id}/certificat-pdf/",
        })
    return resultats


def _lister_certificats(request):
    """Agrège les certificats de tous les modules (incendie, extincteur —
    qui couvre aussi l'éclairage d'urgence via le certificat unifié) en une
    seule liste triée, filtrée par organisation et par les paramètres de
    requête communs (recherche, type, statut, client)."""
    organisation = request.user.organisation

    resultats = []
    type_filtre = request.query_params.get("type")
    if type_filtre in (None, "", "incendie"):
        resultats += _certificats_incendie(organisation)
    if type_filtre in (None, "", "extincteur"):
        resultats += _certificats_extincteur(organisation)

    recherche = (request.query_params.get("recherche") or "").strip().lower()
    if recherche:
        resultats = [
            r for r in resultats
            if recherche in r["adresse"].lower()
            or recherche in r["client_nom"].lower()
            or recherche in r["numero"].lower()
        ]

    statut_filtre = request.query_params.get("statut")
    if statut_filtre == "envoye":
        resultats = [r for r in resultats if r["certificat_envoye"]]
    elif statut_filtre == "non_envoye":
        resultats = [r for r in resultats if not r["certificat_envoye"]]

    conformite_filtre = request.query_params.get("conforme")
    if conformite_filtre == "oui":
        resultats = [r for r in resultats if r["conforme"]]
    elif conformite_filtre == "non":
        resultats = [r for r in resultats if not r["conforme"]]

    client_id = request.query_params.get("client")
    if client_id:
        resultats = [r for r in resultats if str(r["client_id"]) == str(client_id)]

    resultats.sort(key=lambda r: r["date_emission"], reverse=True)
    return resultats


class CertificatsUnifiesView(APIView):
    """Vue agrégée de tous les certificats émis (tous modules confondus) —
    pour le superviseur qui veut retrouver/exporter un certificat sans
    naviguer dans chaque rapport individuellement."""

    permission_classes = [permissions.IsAuthenticated, EstSuperviseur]

    def get(self, request):
        resultats = _lister_certificats(request)
        return Response([
            {**r, "date_emission": r["date_emission"].isoformat()}
            for r in resultats
        ])


class CertificatsExcelView(APIView):
    permission_classes = [permissions.IsAuthenticated, EstSuperviseur]

    def get(self, request):
        resultats = _lister_certificats(request)

        wb = excel_workbook()
        ws = wb.active
        ws.title = "Certificats"
        colonnes = ["Numéro", "Type", "Adresse", "Client", "Date d'émission", "Conforme", "Envoyé au citoyen"]
        excel_ligne_entetes(ws, colonnes, ligne=1)
        for i, r in enumerate(resultats, start=2):
            valeurs = [
                r["numero"], r["type_display"], r["adresse"], r["client_nom"],
                _date_fr(r["date_emission"]),
                "Oui" if r["conforme"] else "Non",
                "Oui" if r["certificat_envoye"] else "Non",
            ]
            for col, valeur in enumerate(valeurs, start=1):
                ws.cell(row=i, column=col, value=valeur)
        excel_ajuster_largeurs(ws, colonnes)
        ws.freeze_panes = "A2"

        return excel_reponse(wb, "Certificats")


# ── Appels de service ───────────────────────────────────────────────────────

class AppelServiceViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]

    def get_serializer_class(self):
        if self.action == "list":
            return AppelServiceListSerializer
        if self.action == "create":
            return AppelServiceCreateSerializer
        return AppelServiceDetailSerializer

    def get_queryset(self):
        user = self.request.user
        qs = AppelService.objects.select_related(
            "batiment", "batiment__client", "cree_par"
        ).prefetch_related("techniciens", "historique").filter(batiment__client__organisation=user.organisation)

        if user.est_technicien():
            qs = qs.filter(techniciens=user)
        # le superviseur voit tout

        statut = self.request.query_params.get("statut")
        client_id = self.request.query_params.get("client")
        if statut:
            qs = qs.filter(statut=statut)
        if client_id:
            qs = qs.filter(batiment__client_id=client_id)

        return qs.distinct()

    def get_permissions(self):
        if self.action in ["create", "destroy", "assigner", "resynchroniser", "terminer_manuellement"]:
            return [permissions.IsAuthenticated(), EstSuperviseur()]
        if self.action in ["demarrer"]:
            return [permissions.IsAuthenticated(), EstSuperviseurOuTechnicien()]
        return super().get_permissions()

    def perform_create(self, serializer):
        appel = serializer.save(cree_par=self.request.user)
        if appel.techniciens.exists():
            from django.utils import timezone

            appel.statut = AppelService.Statut.ASSIGNE
            appel.date_assignation = timezone.now()
            appel.save()
        appel.historiser(self.request.user, "Appel de service créé")
        appel.synchroniser_vers_pubms()

    @action(detail=True, methods=["patch"])
    def assigner(self, request, pk=None):
        """Superviseur seulement — réassigne les techniciens après création.
        Ne re-synchronise PAS pubms (hors scope v1)."""
        appel = self.get_object()
        if "techniciens" not in request.data:
            return Response({"error": "techniciens requis."}, status=status.HTTP_400_BAD_REQUEST)

        appel.techniciens.set(request.data.get("techniciens") or [])
        if appel.techniciens.exists() and appel.statut == AppelService.Statut.OUVERT:
            from django.utils import timezone

            appel.statut = AppelService.Statut.ASSIGNE
            appel.date_assignation = timezone.now()
            appel.save()
        appel.historiser(request.user, "Techniciens réassignés (non répercuté dans pubms)")
        return Response(AppelServiceDetailSerializer(appel).data)

    @action(detail=True, methods=["post"])
    def demarrer(self, request, pk=None):
        from django.utils import timezone

        appel = self.get_object()
        if appel.statut != AppelService.Statut.ASSIGNE:
            return Response({"error": "Doit être assigné avant de démarrer."}, status=status.HTTP_400_BAD_REQUEST)
        appel.statut = AppelService.Statut.EN_COURS
        appel.date_debut = timezone.now()
        appel.save()
        appel.historiser(request.user, "Travail démarré (local, non synchronisé avec pubms)")
        return Response(AppelServiceDetailSerializer(appel).data)

    @action(detail=True, methods=["post"], url_path="terminer-manuellement")
    def terminer_manuellement(self, request, pk=None):
        appel = self.get_object()
        if appel.statut == AppelService.Statut.TERMINE:
            return Response({"error": "Déjà terminé."}, status=status.HTTP_400_BAD_REQUEST)
        appel.terminer_manuellement(request.user)
        return Response(AppelServiceDetailSerializer(appel).data)

    @action(detail=True, methods=["post"])
    def resynchroniser(self, request, pk=None):
        appel = self.get_object()
        appel.synchroniser_vers_pubms()
        return Response(AppelServiceDetailSerializer(appel).data)


class ServiceKeyPermission(permissions.BasePermission):
    """Auth service-à-service par secret partagé — pour les appels serveur-à-serveur (ex: pubms)."""

    def has_permission(self, request, view):
        import secrets
        from django.conf import settings

        key = request.headers.get("X-Service-Key", "")
        expected = settings.INCENDIE_TO_PUBMS_KEY
        return bool(key) and bool(expected) and secrets.compare_digest(key, expected)


class PubmsCallbackAppelServiceView(APIView):
    """Reçoit la notification de pubms qu'une Tache liée à un appel de service est terminée."""

    authentication_classes = []
    permission_classes = [ServiceKeyPermission]

    def post(self, request):
        numero = request.data.get("appel_numero")
        appel_id = request.data.get("appel_id")
        nouveau_statut = request.data.get("statut")

        try:
            if numero:
                appel = AppelService.objects.get(numero=numero)
            else:
                appel = AppelService.objects.get(id=appel_id)
        except AppelService.DoesNotExist:
            return Response({"error": "Appel de service introuvable."}, status=status.HTTP_404_NOT_FOUND)

        if nouveau_statut != "termine":
            return Response({"message": "Ignoré — seul le statut 'termine' déclenche une fermeture."}, status=status.HTTP_200_OK)

        if appel.statut == AppelService.Statut.TERMINE:
            return Response({"message": "Déjà fermé.", "statut": appel.statut}, status=status.HTTP_200_OK)

        appel.terminer_depuis_pubms()
        return Response({"message": f"{appel.numero} marqué terminé.", "statut": appel.statut}, status=status.HTTP_200_OK)